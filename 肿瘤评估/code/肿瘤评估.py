#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse

keys1list = [r'{change}', '[Subject]', '[TMASPD]']
keys2list = [r'{change}', '[Subject]', '[InstanceName]', '[TMASPD]', '[TMADAT]']
keys3list = [r'{change}', '[Subject]', '[InstanceName]', '[TMADAT]', '[TMASTAT]']
keys4list = [r'{change}', '[Subject]', '[InstanceName]', '[IOTATL]', '[IOTATLT]', '[IOTANTL]', '[IOTANTLT]']


def findkeyscolumn(ws, keyslist):
    keys = {}
    for column in range(1, ws.max_column+1):
        row_letter = get_column_letter(column)
        for key in keyslist:
            if key in ws[row_letter+'1'].value:
                keys.setdefault(key, row_letter)
                keyslist.remove(key)
                break
    return keys


def mark(ws, col, row, msg):
    ws[col+str(row)] = msg
    return


def message(msg, rsg):
    if msg == '':
        msg = rsg
    else:
        msg = '\n'.join([msg, rsg])
    return msg


def data1(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        TMASPD = set()
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue

        Subject = ws[keys['[Subject]']+str(row)].value
        TMASPD.add(ws[keys['[TMASPD]']+str(row)].value)
        TMASPD = list(TMASPD)

        data_ws.setdefault(Subject, TMASPD[0])
    return data_ws


def data2(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        TMASPD = set()
        InstanceName = set()
        TMADAT = set()
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        if ws[keys['[TMADAT]']+str(row)].value == None:
            continue

        Subject = ws[keys['[Subject]']+str(row)].value
        InstanceName.add(ws[keys['[InstanceName]']+str(row)].value)
        TMASPD.add(ws[keys['[TMASPD]']+str(row)].value)
        TMADAT.add(ws[keys['[TMADAT]']+str(row)].value)
        
        InstanceName = list(InstanceName)
        TMASPD = list(TMASPD)
        TMADAT = list(TMADAT)

        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(InstanceName[0], {'TMASPD':TMASPD[0], 'TMADAT':TMADAT[0]})
    return data_ws


def data3(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        InstanceName = set()
        TMADAT = set()
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        if ws[keys['[TMADAT]']+str(row)].value == None:
            continue 

        Subject = ws[keys['[Subject]']+str(row)].value
        InstanceName.add(ws[keys['[InstanceName]']+str(row)].value)
        TMADAT.add(ws[keys['[TMADAT]']+str(row)].value)
        TMASTAT = ws[keys['[TMASTAT]']+str(row)].value

        InstanceName = list(InstanceName)
        TMADAT = list(TMADAT)

        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(InstanceName[0], {'TMADAT':TMADAT[0], 'TMASTAT':{TMASTAT}})
        if data_ws[Subject][InstanceName[0]] != InstanceName[0]:
            data_ws[Subject][InstanceName[0]]['TMASTAT'].add(TMASTAT)
    return data_ws


def data4(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        Subject = ws[keys['[Subject]']+str(row)].value
        InstanceName = ws[keys['[InstanceName]']+str(row)].value
        IOTATL = ws[keys['[IOTATL]']+str(row)].value
        IOTATLT = ws[keys['[IOTATLT]']+str(row)].value
        IOTANTL = ws[keys['[IOTANTL]']+str(row)].value
        IOTANTLT = ws[keys['[IOTANTLT]']+str(row)].value

        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(row, {'InstanceName':InstanceName, 'IOTATL':IOTATL, 'IOTATLT':IOTATLT, 'IOTANTL':IOTANTL, 'IOTANTLT':IOTANTLT})
    return data_ws


def bbzcheck(data_ws1, data_ws2, data_ws4, ws4):
    ws4.insert_cols(1)
    ws4['A1'].value = '靶病灶检查结果'

    for id in data_ws4:
        pid = data_ws4[id]
        for row in pid:
            msg = ''
            rpid = pid[row]
            cont = True
            if id in data_ws2:
                pid_ws2 = data_ws2[id]
                Instancelist = list()
                TMASPDlist = list()
                sorted_pid_ws2 = sorted(pid_ws2.items(), key = lambda time:time[1]['TMADAT'])
                for i in range(0, len(sorted_pid_ws2)):
                    Instancelist.append(sorted_pid_ws2[i][0])
                    TMASPDlist.append(sorted_pid_ws2[i][1]['TMASPD'])                                    
                InstanceName = rpid['InstanceName']
                IOTATL = rpid['IOTATL']
                IOTATLT = rpid['IOTATLT']

                if id in data_ws1:
                    TMASPD_raw = data_ws1[id]
                else:
                    if 'NE' in IOTATL or 'NA' in IOTATL:
                        if IOTATLT:
                            rsg = 'Info:该行无靶病灶患者筛选期信息，内容为NE/NA，存在说明'
                        else:
                            rsg = 'Error:该行无靶病灶患者筛选期信息，内容为NE/NA，但说明缺失'
                    else:
                        rsg = 'Error:该患者无靶病灶筛选期对应信息'
                    cont = False

                if InstanceName in Instancelist and cont:
                    index = Instancelist.index(InstanceName)
                    check = TMASPDlist[index]
                    checklist = TMASPDlist[:index+1]
                    checklist.append(TMASPD_raw)
                    TMASPD_min = min(checklist)
                    zerocheck = True
                    if TMASPD_min != 0:
                        PRcheck = (check - TMASPD_raw)/TMASPD_raw
                        PDcheck = (check - TMASPD_min)/TMASPD_min
                    else:
                        zerocheck = False
                        
                    if zerocheck:    
                        if PDcheck >= 0.2 and abs(check - TMASPD_min) >= 5:
                            result = 'PD'
                        elif abs(PRcheck) >= 0.3:
                            result = 'PR'
                        else:
                            result = 'SD'

                    if 'NE' in IOTATL or 'NA' in IOTATL:
                        if IOTATLT:
                            rsg = 'Info:该行靶病灶结果为NE/NA，存在说明'
                        else:
                            rsg = 'Error:该行靶病灶结果为NE/NA，但说明缺失'
                    elif result in IOTATL:
                        rsg = 'Info:该行结果与靶病灶匹配成功'
                    elif not zerocheck:
                        rsg = 'Warn:此次检测数值为零，需提供说明'
                    else:
                        rsg = 'Error:该行靶病灶结果应为{}，与本行匹配失败'.format(result)
                else:
                    if 'NE' in IOTATL or 'NA' in IOTATL:
                        if IOTATLT:
                            rsg = 'Info:该行无靶病灶访视信息，内容为NE/NA，存在说明'
                        else:
                            rsg = 'Error:该行无靶病灶访视信息，内容为NE/NA，但说明缺失'
                    else:                        
                        rsg = 'Error:该患者在靶病灶页面无访视 {} 信息'.format(InstanceName)
            else:
                if 'NE' in IOTATL or 'NA' in IOTATL:
                    if IOTATLT:
                        rsg = 'Info:该行无靶病灶患者信息，内容为NE/NA，存在说明'
                    else:
                        rsg = 'Error:该行无靶病灶患者信息，内容为NE/NA，但说明缺失'
                else:
                    rsg = 'Error:该患者无靶病灶对应信息'
            msg = message(msg, rsg)            
            mark(ws4, 'A', row, msg)
    return ws4


def fbbzcheck(data_ws3, data_ws4, ws4):
    ws4.insert_cols(1)
    ws4['A1'].value = '非靶病灶检查结果'

    for id in data_ws4:
        pid = data_ws4[id]
        for row in pid:
            msg = ''
            rpid = pid[row]
            InstanceName = rpid['InstanceName']
            IOTANTL = rpid['IOTANTL']
            IOTANTLT = rpid['IOTANTLT']
            if id in data_ws3:
                pid_ws2 = data_ws3[id]
                if InstanceName in pid_ws2:
                    ipid_ws2 = pid_ws2[InstanceName]
                    if '明显进展' in ipid_ws2['TMASTAT']:
                        result = 'PD'
                    elif '存在' in ipid_ws2['TMASTAT']:
                        result = 'Non-CR'
                    elif len(ipid_ws2['TMASTAT'].union({'消失'})) == 1:
                        result = 'CR'
                    
                    if 'NE' in IOTANTL or 'NA' in IOTANTL:
                        if IOTANTLT:
                            rsg = 'Info:该行非靶病灶结果为NE/NA，存在说明'
                        else:
                            rsg = 'Error:该行非靶病灶结果为NE/NA，但说明缺失'
                    elif result in IOTANTL:
                        rsg = 'Info:该行结果与非靶病灶匹配成功'
                    else:
                        rsg = 'Error:该行非靶病灶结果应为{}，与本行匹配失败'.format(result)
                else:
                    if 'NE' in IOTANTL or 'NA' in IOTANTL:
                        if IOTANTLT:
                            rsg = 'Info:该行无非靶病灶访视信息，内容为NE/NA，存在说明'
                        else:
                            rsg = 'Error:该行无非靶病灶访视信息，内容为NE/NA，但说明缺失'
                    else:                       
                        rsg = 'Error:该患者在非靶病灶页面无访视 {} 信息'.format(InstanceName)
            else:
                if 'NNT' in IOTANTL:
                    rsg = 'Info:该患者无非靶病灶'
                elif 'NE' in IOTANTL or 'NA' in IOTANTL:
                        if IOTANTLT:
                            rsg = 'Info:该行无非靶病灶患者信息，内容为NE/NA，存在说明'
                        else:
                            rsg = 'Error:该行无非靶病灶患者信息，内容为NE/NA，但说明缺失'
                else:
                    rsg = 'Error:该患者无非靶病灶对应信息'
            msg = message(msg, rsg)            
            mark(ws4, 'A', row, msg)
    return ws4


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--cancer", default=r'C:\Users\yao.dong\Desktop\CIBI308C301本地\project4\20210728\CIBI308C303_肿瘤评估_20210728.xlsx', help="Please add AE file full path")
    parser.add_argument("--sxq", default=r'TMA001_1|肿瘤评估-靶病灶-筛选期', help="Please set sheet name of cb")
    parser.add_argument("--bbz", default=r'TMA001_2|肿瘤评估-靶病灶', help="Please set sheet name of ae")
    parser.add_argument("--fbbz", default=r'TMA001_4|肿瘤评估-非靶病灶', help="Please set sheet name of cb")
    parser.add_argument("--recist", default=r'IOTA001_1|实体瘤疗效评估（RECIST v1.1）', help="Please set sheet name of cb")
    parser.add_argument("--flow", default="all", help="Please state the flow you need to run")

    args = parser.parse_args()

    cancer_path = args.cancer
    sxq_sheet = args.sxq
    bbz_sheet = args.bbz
    fbbz_sheet = args.fbbz
    recist_sheet = args.recist
    flow = args.flow

    cancer_pathlist = cancer_path.split('.')

    wbsavepath = '.'.join([''.join([cancer_pathlist[0], '_checkout']), cancer_pathlist[1]])
    try:
        wb = openpyxl.load_workbook(cancer_path)
        ws1 = wb[sxq_sheet]
        ws2 = wb[bbz_sheet]
        ws3 = wb[fbbz_sheet]
        ws4 = wb[recist_sheet]
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)
        keys3 = findkeyscolumn(ws3, keys3list)
        keys4 = findkeyscolumn(ws4, keys4list)

        data_ws1 = data1(ws1, keys1)
        data_ws2 = data2(ws2, keys2)
        data_ws3 = data3(ws3, keys3)
        data_ws4 = data4(ws4, keys4)
        
        ws4 = fbbzcheck(data_ws3, data_ws4, ws4)
        ws4 = bbzcheck(data_ws1, data_ws2, data_ws4, ws4)

        wb.save(wbsavepath)

    finally:
        wb.close()