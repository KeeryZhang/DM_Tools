#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse

keys1list = [r'{change}', '[Subject]', '[TMASPD]']
keys2list = [r'{change}', '[Subject]', '[InstanceName]', '[TMASPD]', '[TMADAT]']
keys3list = [r'{change}', '[Subject]', '[InstanceName]', '[TMADAT]', '[TMASTAT]']
keys4list = [r'{change}', '[Subject]', '[InstanceName]', '[IOTATL]', '[IOTATLT]', '[IOTANTL]', '[IOTANTLT]']


def findkeyscolumn(ws, keyslist):
    keys = {}
    for column in range(1, ws.max_column+1):
        row_letter = get_column_letter(column)
        for key in keyslist:
            if key in ws[row_letter+'1'].value:
                keys.setdefault(key, row_letter)
                keyslist.remove(key)
                break
    return keys


def mark(ws, col, row, msg):
    ws[col+str(row)] = msg
    return


def message(msg, rsg):
    if msg == '':
        msg = rsg
    else:
        msg = '\n'.join([msg, rsg])
    return msg


def data1(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        TMASPD = set()
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue

        Subject = ws[keys['[Subject]']+str(row)].value
        TMASPD.add(ws[keys['[TMASPD]']+str(row)].value)
        TMASPD = list(TMASPD)

        data_ws.setdefault(Subject, TMASPD[0])
    return data_ws


def data2(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        TMASPD = set()
        InstanceName = set()
        TMADAT = set()
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        if ws[keys['[TMADAT]']+str(row)].value == None:
            continue

        Subject = ws[keys['[Subject]']+str(row)].value
        InstanceName.add(ws[keys['[InstanceName]']+str(row)].value)
        TMASPD.add(ws[keys['[TMASPD]']+str(row)].value)
        TMADAT.add(ws[keys['[TMADAT]']+str(row)].value)
        
        InstanceName = list(InstanceName)
        TMASPD = list(TMASPD)
        TMADAT = list(TMADAT)

        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(InstanceName[0], {'TMASPD':TMASPD[0], 'TMADAT':TMADAT[0], 'PR':0, 'PD':0})
    return data_ws


def data3(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        InstanceName = set()
        TMADAT = set()
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        if ws[keys['[TMADAT]']+str(row)].value == None:
            continue 

        Subject = ws[keys['[Subject]']+str(row)].value
        InstanceName.add(ws[keys['[InstanceName]']+str(row)].value)
        TMADAT.add(ws[keys['[TMADAT]']+str(row)].value)
        TMASTAT = ws[keys['[TMASTAT]']+str(row)].value

        InstanceName = list(InstanceName)
        TMADAT = list(TMADAT)

        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(InstanceName[0], {'TMADAT':TMADAT[0], 'TMASTAT':{TMASTAT}})
        if data_ws[Subject][InstanceName[0]] != InstanceName[0]:
            data_ws[Subject][InstanceName[0]]['TMASTAT'].add(TMASTAT)
    return data_ws


def data4(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        Subject = ws[keys['[Subject]']+str(row)].value
        InstanceName = ws[keys['[InstanceName]']+str(row)].value
        IOTATL = ws[keys['[IOTATL]']+str(row)].value
        IOTATLT = ws[keys['[IOTATLT]']+str(row)].value
        IOTANTL = ws[keys['[IOTANTL]']+str(row)].value
        IOTANTLT = ws[keys['[IOTANTLT]']+str(row)].value

        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(row, {'InstanceName':InstanceName, 'IOTATL':IOTATL, 'IOTATLT':IOTATLT, 'IOTANTL':IOTANTL, 'IOTANTLT':IOTANTLT, 'order':0})
    return data_ws


def exist(checkitem, container):
    if checkitem in container:
        return True
    else:
        return False


def bbzresult(InstanceName, crossbase, crosschecklist, crossinstancelist):
    index = crossinstancelist.index(InstanceName)
    check = crosschecklist[index]
    checklist = crosschecklist[:index+1]
    checklist.append(crossbase)
    TMASPD_min = min(checklist)
    zerocheck = True
    if TMASPD_min != 0:
        PRcheck = (check - crossbase)/crossbase
        PDcheck = (check - TMASPD_min)/TMASPD_min
    else:
        zerocheck = False
        
    if zerocheck:    
        if PDcheck >= 0.2 and abs(check - TMASPD_min) >= 5:
            result = 'PD'
        elif abs(PRcheck) >= 0.3:
            result = 'PR'
        else:
            result = 'SD'
    else:
        result = '0'
    return result


def newbbz(data_ws1, data_ws2, data_ws4, ws4):
    ws4.insert_cols(1)
    ws4['A1'].value = '?????????????????????'

    for id in data_ws4:
        pid = data_ws4[id]
        idcheck = exist(id, data_ws2)
        IOTATLlist = list()
        NAcheck = False
        crossbase = None
        crosschecklist = None
        crossinstancelist = None

        if idcheck:
            pid_ws2 = data_ws2[id]
            Instancelist = list()
            TMASPDlist = list()
            sorted_pid_ws2 = sorted(pid_ws2.items(), key = lambda time:time[1]['TMADAT'])
            for i in range(0, len(sorted_pid_ws2)):
                Instancelist.append(sorted_pid_ws2[i][0])
                TMASPDlist.append(sorted_pid_ws2[i][1]['TMASPD'])     
                                               
            for row in pid:
                rpid = pid[row]
                InstanceName = rpid['InstanceName']
                for index in range(0,len(Instancelist)):
                    if InstanceName == Instancelist[index]:
                        rpid['order'] = index
            
            rowlist = sorted(pid.items(), key = lambda order:order[1]['order'])

        for i in range(0,len(rowlist)):
            row = rowlist[i][0]
            rpid = rowlist[i][1]
            msg = ''
            cont = True
            InstanceName = rpid['InstanceName']
            IOTATL = rpid['IOTATL']
            IOTATLT = rpid['IOTATLT']
            if idcheck:
                if id in data_ws1:
                    TMASPD_raw = data_ws1[id]
                else:
                    if 'NE' in IOTATL or 'NA' in IOTATL:
                        if IOTATLT:
                            rsg = 'Info:???????????????????????????????????????????????????NE/NA???????????????'
                        else:
                            rsg = 'Error:???????????????????????????????????????????????????NE/NA??????????????????'
                    else:
                        rsg = 'Error:??????????????????????????????????????????'
                    cont = False
                
                if not NAcheck:
                    crossbase = TMASPD_raw
                    crosschecklist = TMASPDlist
                    crossinstancelist = Instancelist

                if InstanceName in Instancelist and cont:
                    result = bbzresult(InstanceName, crossbase, crosschecklist, crossinstancelist)

                    if 'NE' in IOTATL or 'NA' in IOTATL:
                        if '??????' in IOTATLT or '????????????' in IOTATLT:    
                            NAcheck = True                        
                            # for r in rowlist:
                            #     IOTATLlist.append(r[1]['IOTATL'])
                            # for index in range((Instancelist.index(InstanceName)-len(Instancelist)-1),(-len(Instancelist)),-1):
                            #     if 'PD' in IOTATLlist[index]:
                            #         break
                            index = crossinstancelist.index(InstanceName)
                            crossbase = TMASPDlist[index]
                            crosschecklist = TMASPDlist[(Instancelist.index(InstanceName)+1):]
                            crossinstancelist = Instancelist[(Instancelist.index(InstanceName)+1):]
                            rsg = 'Info:????????????????????????NA????????????????????????'
                        elif IOTATLT:
                            rsg = 'Info:????????????????????????NE/NA???????????????'
                        elif IOTATLT == None:
                            rsg = 'Error:????????????????????????NE/NA??????????????????'
                    elif result in IOTATL:
                        rsg = 'Info:????????????????????????????????????'
                    elif result == '0':
                        rsg = 'Warn:??????????????????????????????????????????'
                    else:
                        rsg = 'Error:???????????????????????????{}????????????????????????'.format(result)
                else:
                    if 'NE' in IOTATL or 'NA' in IOTATL:
                        if IOTATLT:
                            rsg = 'Info:??????????????????????????????????????????NE/NA???????????????'
                        else:
                            rsg = 'Error:??????????????????????????????????????????NE/NA??????????????????'
                    else:                        
                        rsg = 'Error:???????????????????????????????????? {} ??????'.format(InstanceName)

            else:
                if 'NE' in IOTATL or 'NA' in IOTATL:
                    if IOTATLT:
                        rsg = 'Info:??????????????????????????????????????????NE/NA???????????????'
                    else:
                        rsg = 'Error:??????????????????????????????????????????NE/NA??????????????????'
                else:
                    rsg = 'Error:?????????????????????????????????'                
            msg = message(msg, rsg)            
            mark(ws4, 'A', row, msg)
    return ws4

def fbbzcheck(data_ws3, data_ws4, ws4):
    ws4.insert_cols(1)
    ws4['A1'].value = '????????????????????????'

    for id in data_ws4:
        pid = data_ws4[id]
        for row in pid:
            msg = ''
            rpid = pid[row]
            InstanceName = rpid['InstanceName']
            IOTANTL = rpid['IOTANTL']
            IOTANTLT = rpid['IOTANTLT']
            if id in data_ws3:
                pid_ws2 = data_ws3[id]
                if InstanceName in pid_ws2:
                    ipid_ws2 = pid_ws2[InstanceName]
                    if '????????????' in ipid_ws2['TMASTAT']:
                        result = 'PD'
                    elif '??????' in ipid_ws2['TMASTAT']:
                        result = 'Non-CR'
                    elif len(ipid_ws2['TMASTAT'].union({'??????'})) == 1:
                        result = 'CR'
                    
                    if 'NE' in IOTANTL or 'NA' in IOTANTL:
                        if IOTANTLT:
                            rsg = 'Info:???????????????????????????NE/NA???????????????'
                        else:
                            rsg = 'Error:???????????????????????????NE/NA??????????????????'
                    elif result in IOTANTL:
                        rsg = 'Info:???????????????????????????????????????'
                    else:
                        rsg = 'Error:??????????????????????????????{}????????????????????????'.format(result)
                else:
                    if 'NE' in IOTANTL or 'NA' in IOTANTL:
                        if IOTANTLT:
                            rsg = 'Info:?????????????????????????????????????????????NE/NA???????????????'
                        else:
                            rsg = 'Error:?????????????????????????????????????????????NE/NA??????????????????'
                    else:                       
                        rsg = 'Error:??????????????????????????????????????? {} ??????'.format(InstanceName)
            else:
                if 'NNT' in IOTANTL:
                    rsg = 'Info:????????????????????????'
                elif 'NE' in IOTANTL or 'NA' in IOTANTL:
                        if IOTANTLT:
                            rsg = 'Info:?????????????????????????????????????????????NE/NA???????????????'
                        else:
                            rsg = 'Error:?????????????????????????????????????????????NE/NA??????????????????'
                else:
                    rsg = 'Error:????????????????????????????????????'
            msg = message(msg, rsg)            
            mark(ws4, 'A', row, msg)
    return ws4


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--cancer", default=r'C:\Users\yao.dong\Desktop\CIBI308C301??????\project4\????????????\20210728\CIBI308C303_????????????_20210728.xlsx', help="Please add AE file full path")
    parser.add_argument("--sxq", default=r'TMA001_1|????????????-?????????-?????????', help="Please set sheet name of cb")
    parser.add_argument("--bbz", default=r'TMA001_2|????????????-?????????', help="Please set sheet name of ae")
    parser.add_argument("--fbbz", default=r'TMA001_4|????????????-????????????', help="Please set sheet name of cb")
    parser.add_argument("--recist", default=r'IOTA001_1|????????????????????????RECIST v1.1???', help="Please set sheet name of cb")
    parser.add_argument("--flow", default="all", help="Please state the flow you need to run")

    args = parser.parse_args()

    cancer_path = args.cancer
    sxq_sheet = args.sxq
    bbz_sheet = args.bbz
    fbbz_sheet = args.fbbz
    recist_sheet = args.recist
    flow = args.flow

    cancer_pathlist = cancer_path.split('.xlsx')

    wbsavepath = ''.join([''.join([cancer_pathlist[0], '_checkout']), '.xlsx'])
    try:
        wb = openpyxl.load_workbook(cancer_path)
        ws1 = wb[sxq_sheet]
        ws2 = wb[bbz_sheet]
        ws3 = wb[fbbz_sheet]
        ws4 = wb[recist_sheet]
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)
        keys3 = findkeyscolumn(ws3, keys3list)
        keys4 = findkeyscolumn(ws4, keys4list)

        data_ws1 = data1(ws1, keys1)
        data_ws2 = data2(ws2, keys2)
        data_ws3 = data3(ws3, keys3)
        data_ws4 = data4(ws4, keys4)
        
        ws4 = fbbzcheck(data_ws3, data_ws4, ws4)
        ws4 = newbbz(data_ws1, data_ws2, data_ws4, ws4)

        wb.save(wbsavepath)

    finally:
        wb.close()