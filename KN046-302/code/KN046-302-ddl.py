#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message

keyslist = ['[Subject]', '[SiteNumber]']

SHEETS_PATH = "..\sheets"

def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file

if __name__ == "__main__":
    files = get_files()
    ddl = get_a_file(files, "KN046-302_datadumplistings")
    ddl_path = os.path.join(SHEETS_PATH, ddl)
    ddl_pathlist = ddl_path.split('.xlsx')
    wbsavepath = ''.join([''.join([ddl_pathlist[0], '_checkout']), '.xlsx'])

    try:
        wb = openpyxl.load_workbook(ddl_path)
        sheets_list = deepcopy(wb.sheetnames)
        sheets_list.remove('SourceSummary')
        sheets_list.remove('RowChangeSummary')

        for sheet in sheets_list:
            ws = wb[sheet]
            keyslist_tmp = deepcopy(keyslist)
            keys = findkeyscolumn(ws, keyslist_tmp)

            for row in range(2, ws.max_row+1):
                subject = str(ws[keys['[Subject]']+str(row)].value)
                if subject == 'None':
                    continue
                site = subject[:3]
                ws[keys['[SiteNumber]']+str(row)] = site

        wb.save(wbsavepath)

    finally:
        wb.close()