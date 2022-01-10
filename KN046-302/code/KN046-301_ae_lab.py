#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from datetime import datetime
import os
import sys
import argparse
import copy
from copy import deepcopy
import pdb

sys.path.append("..\..")
from tool_lib.outs.Codelist_map import Codelist_map as Codelist_map
from tool_lib.outs.Coderevert_map import Coderevert_map_withsymbol as Coderevert_map

from tool_lib.utils import findkeyscolumn, parse_dmy, mark


keys1list = [r'{change}','[Subject]','[AEYN]','AETERM_DECOD','[AESTDAT]',
             '[AEENDAT]','[AEONGO]']
keys2list = [r'{change}','[Subject]','[RecordDate]','[AnalyteName]',
             '[AnalyteValue]','[LabFlag]','[ClinSigValue]','[InstanceName]',
             '[ClinSigComment]']

SHEETS_PATH = "..\sheets"

def data1(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        if ws[keys[r'{change}']+str(row)] == 'deleted' or \
           ws[keys['[AEYN]']+str(row)] == '否' or \
           ws[keys['[AEYN]']+str(row)] == None:
            continue
        patientId = ws[keys['[Subject]']+str(row)].value
        AE_PT = ws[keys['AETERM_DECOD']+str(row)].value
        AEONGO = ws[keys['[AEONGO]']+str(row)].value

        ST = ws[keys['[AESTDAT]']+str(row)].value
        EN = ws[keys['[AEENDAT]']+str(row)].value
        
        AnalyteNamelist = list()
        AE_PTlist = list()
        if AE_PT in Coderevert_map:
            for i in Coderevert_map[AE_PT]:
                AnalyteNamelist.append(i[0])

            for AE in Codelist_map:
                if AE[0] in AnalyteNamelist:
                    AE_PTlist.extend([str(x) for x in Codelist_map[AE]])
            AE_PTlist=list(set(AE_PTlist))
            AE_PTlist.sort()
            AE_PTlist = tuple(AE_PTlist)
        else:
            AE_PTlist.append(AE_PT)
            AE_PTlist = tuple(AE_PTlist)
        data_ws.setdefault(patientId, {})
        data_ws[patientId].setdefault(AE_PTlist, {})
        data_ws[patientId][AE_PTlist].setdefault(ST, {})
        data_ws[patientId][AE_PTlist][ST]. \
            setdefault(row, {'EN':EN, 'AEONGO':AEONGO})
    return data_ws


def data2(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        patientId = ws[keys['[Subject]']+str(row)].value
        RecordDate = ws[keys['[RecordDate]']+str(row)].value
        AnalyteName = ws[keys['[AnalyteName]']+str(row)].value
        AnalyteValue = ws[keys['[AnalyteValue]']+str(row)].value
        LabFlag = ws[keys['[LabFlag]']+str(row)].value
        CS = ws[keys['[ClinSigValue]']+str(row)].value
        visitname = ws[keys['[InstanceName]']+str(row)].value
        ClinSigComment = ws[keys['[ClinSigComment]']+str(row)].value

        data_ws.setdefault(patientId, {})
        data_ws[patientId].setdefault(AnalyteName, {})
        data_ws[patientId][AnalyteName]. \
            setdefault(row, {
                'AnalyteName': AnalyteName,
                'RecordDate': RecordDate,
                'AnalyteValue': AnalyteValue,
                'LabFlag': LabFlag,
                'CS': CS,
                'visitname': visitname,
                'ClinSigComment': ClinSigComment})
    return data_ws


def lab2ae(data_ws1, data_ws2, ws2):
    ws2.insert_cols(1)
    ws2['A1'].value = 'YD Comment'

    for id in data_ws2:
        pid = data_ws2[id]
        for AnalyteName in pid:            
            STapid_ws1 = {}
            apid = pid[AnalyteName]
            sortedapid = sorted(apid.items(), 
                key = lambda time:time[1]['RecordDate'])
            AE_event = False
            raw_ws1 = None
            AETERM_PT = None
            for row_ws2 in sortedapid:
                rs = row_ws2[1]
                if rs['LabFlag'] in ['+', '-'] and rs['CS'] == '异常有临床意义':
                    if id not in data_ws1.keys():
                        msg = 'Error:该患者在AE页面无记录'
                        mark(ws2, 'A', row_ws2[0], msg)
                        continue       
                    
                    if (AnalyteName, rs['LabFlag']) not in Codelist_map:
                        msg = 'Warn:该行分析物名称无对应不良事件名称，请核查'
                        mark(ws2, 'A', row_ws2[0], msg)
                        continue 

                    if not AE_event:
                        codelist = Codelist_map[(AnalyteName, rs['LabFlag'])]
                        codecheck = False
                        for code in codelist:
                            for subcode in data_ws1[id]:
                                if code in subcode:
                                    codecheck = True
                                    AETERM_PT = subcode
                                    break
                            if codecheck:
                                break
                        
                        if not codecheck:
                            msg = 'Error:该患者在AE页面无{}记录'.format(AnalyteName)
                            mark(ws2, 'A', row_ws2[0], msg)
                            continue                         
                    
                        apid_ws1 = data_ws1[id][AETERM_PT]
                        if rs['RecordDate'] in apid_ws1:
                            STapid_ws1_pre = copy.deepcopy(
                                apid_ws1[rs['RecordDate']])
                            AE_event = True
                            msg = 'Info:该行为AE开始，在AE页面有相应记录'
                            mark(ws2, 'A', row_ws2[0], msg)
                            STapid_ws1 = sorted(STapid_ws1_pre.items())
                            raw_ws1 = STapid_ws1[0]
                            continue

                        else:
                            msg = 'Error:该行有{0}发生，但未在AE页面找到相应记录' \
                                .format(AnalyteName)
                            mark(ws2, 'A', row_ws2[0], msg)
                            continue
                                        
                    else:
                        checkline = False
                    #     for line in STapid_ws1:
                    #         if rs['RecordDate'] == line[1]['GR']:
                    #             raw_ws1 = line
                    #             msg = 'Info:该行对应AE页面第{}行，发生级别变化'.format(raw_ws1[0])
                    #             mark(ws2, 'A', row_ws2[0], msg)
                    #             checkline = True
                    #             break
                        if not checkline:
                            msg = 'Info:该行处于第{0}行AE {1} 发生中' \
                                .format(raw_ws1[0], code)
                            mark(ws2, 'A', row_ws2[0], msg)
                            continue

                if rs['LabFlag'] == None and rs['AnalyteValue'] == None:
                    msg = 'Info:该行未录入数据'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == None and \
                     rs['AnalyteValue'] in ['NA', 'ND', 'UN', '不适用', '未做']:
                    msg = 'Info:该项目未检测'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == None:
                    msg = 'Error:实验室检测范围缺失'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == '0' and not AE_event:
                    msg = 'Info:该行数值在正常范围内，无需核查'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == '0' and AE_event:
                    EN = raw_ws1[1]['EN']
                    if rs['RecordDate'] == EN:
                        msg = 'Info:该行为AE结束，在AE页面有相应记录'
                    else:
                        msg = 'Error:该行应为AE结束日期，但在AE页面第{0}行{1}无对应' \
                            .format(raw_ws1[0], code)
                    raw_ws1 = None
                    mark(ws2, 'A', row_ws2[0], msg)
                    AE_event = False
                    continue
                elif rs['LabFlag'] in ['+', '-'] and \
                     rs['CS'] == '异常无临床意义' and \
                     not AE_event:
                    msg = 'Info:该行NCS，无需核查'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] in ['+', '-'] and \
                     rs['CS'] == '异常无临床意义' and \
                     AE_event:
                    EN = raw_ws1[1]['EN']
                    if rs['RecordDate'] == EN:
                        msg = 'Info:该行NCS为AE结束，在AE页面有相应记录'
                    else:
                        msg = 'Error:该行NCS应为AE结束日期，但在AE页面第{}行无对应' \
                            .format(raw_ws1[0])
                    raw_ws1 = None
                    mark(ws2, 'A', row_ws2[0], msg)
                    AE_event = False
                    continue
                else:
                    msg='error:该行异常，但未判断CS or NCS'
                    mark(ws2,'A',row_ws2[0],msg)

    return ws2


def ae2lab(data_ws1, data_ws2, ws1):
    ws1.insert_cols(1)
    ws1['A1'].value = '结束日期比较'
    ws1.insert_cols(1)
    ws1['A1'].value = '开始日期比较'
    
    for id in data_ws1:
        pid = data_ws1[id]
        for AE_PT_list in pid:
            apid = pid[AE_PT_list]
            for AE_PT in AE_PT_list:
                if AE_PT in Coderevert_map:
                    AnalyteNames = Coderevert_map[AE_PT]
                else:
                    AnalyteNames = None
                for ST in apid:                
                    sapid = apid[ST]
                    # sortedsapid = sorted(sapid.items(), key = lambda time:time[1]['GR'])
                    for row_ws1 in sapid:
                        msg = ''
                        msg_s = ''
                        msg_e = ''
                        if id not in data_ws2.keys():
                            msg = 'Error:该患者在Lab页面无记录'
                            mark(ws1, 'A', row_ws1, msg)
                            continue
                        pid_ws2 = data_ws2[id]

                        AnalyteNamelist = list(pid_ws2.keys())
                        AN_left = set()
                        if AnalyteNames is not None:
                            for An_Flag in AnalyteNames:
                                if An_Flag[0] in AnalyteNamelist:
                                    AN_left.add(An_Flag[0])
                        End_lines = list()
                        pid_ws2_list = list()
                        for key in pid_ws2:
                            if AnalyteNames is not None:
                                for AnalyteName in AnalyteNames:
                                    if key == AnalyteName[0]:
                                        pid_ws2_list.append(pid_ws2[key])
                        # pid_ws2_list = list(pid_ws2.values())
                        pid_ws2_all = dict()
                        for i in pid_ws2_list:
                            for key in i:
                                # if i[key]['CS'] == "异常有临床意义":
                                # if key in pid_ws2_all and 
                                pid_ws2_all.setdefault(key, i[key])
                        pid_ws2_sort = sorted(pid_ws2_all.items(),
                            key = lambda time:time[1]['RecordDate'])
                        
                        AnalyteNamecheck = False
                        if AnalyteNames == None:
                            msg = 'Warn:该行不良事件{}无对应分析物，需手动核查'.format(AE_PT)
                            mark(ws1, 'A', row_ws1, msg)
                            mark(ws1, 'B', row_ws1, msg)
                            continue

                        for An_Flag in AnalyteNames:
                            if An_Flag[0] in AnalyteNamelist:
                                AnalyteNamecheck = True
                                # apid_ws2 = pid_ws2[An_Flag[0]]
                                STcheck = False
                                GRcheck = False
                                ENcheck = False
                                GR = True
                                EN = False
                                # if row_ws1[1]['GR'] != datetime(1970, 1, 1):
                                #     GR = True
                                # if (sapid[row_ws1]['EN'] != datetime(1970, 1, 1) or \
                                #     sapid[row_ws1]['EN'] is not None) and \
                                #     sapid[row_ws1]['AEONGO'] != "是":
                                if sapid[row_ws1]['EN'] != datetime(1970, 1, 1):
                                    EN = True
                                for i in pid_ws2_sort:
                                    row_ws2 = i[0]
                                    rapid_ws2 = i[1]
                                    if rapid_ws2['RecordDate'] == ST:
                                        if rapid_ws2['CS'] == '异常有临床意义' and STcheck == False:
                                            STcheck = True
                                            msg_s = ("Info:该AE记录开始日期与Lab页面第{0}行"
                                            "{1} {2} {3} {4} 匹配成功".format(
                                                row_ws2, rapid_ws2['visitname'],
                                                rapid_ws2['AnalyteName'], rapid_ws2['CS'],
                                                rapid_ws2['ClinSigComment']))
                                        elif rapid_ws2['CS'] != '异常有临床意义' and STcheck == False:
                                            msg_s = ('Error:该AE记录在lab页面不是CS,与第'
                                            '{0}行 {1} {2} {3} {4} {5}匹配失败' \
                                                .format(
                                                    row_ws2,
                                                    rapid_ws2['visitname'],
                                                    rapid_ws2['AnalyteName'],
                                                    rapid_ws2['CS'],
                                                    rapid_ws2['ClinSigComment'],
                                                    rapid_ws2['RecordDate']))
                                    elif (EN and 
                                         (rapid_ws2['RecordDate'] == sapid[row_ws1]['EN'] or 
                                            sapid[row_ws1]['EN'] is None) and 
                                         ENcheck == False):
                                        if rapid_ws2['CS'] == '异常有临床意义':
                                            if sapid[row_ws1]['AEONGO'] == "是":
                                                msg_e = "Info: 该AE {} 持续" \
                                                    .format(AE_PT)
                                            elif sapid[row_ws1]['AEONGO'] == "否":
                                                msg_e = ('Error:该AE记录在lab页面为CS,'
                                                '与第{0}行 {1} {2} {3} {4} {5}匹配失败' \
                                                    .format(
                                                        row_ws2, 
                                                        rapid_ws2['visitname'],
                                                        rapid_ws2['AnalyteName'],
                                                        rapid_ws2['CS'],
                                                        rapid_ws2['ClinSigComment'],
                                                        rapid_ws2['RecordDate']))
                                        else:
                                            if sapid[row_ws1]['AEONGO'] == "是":
                                                msg_e = ("Error: 该AE记录在lab页面最近日期的一条不是CS，"
                                                "与第{0}行 {1} {2} {3} {4} {5}匹配失败，"
                                                "但AE页面持续为是".format(
                                                    row_ws2,
                                                    rapid_ws2['visitname'],
                                                    rapid_ws2['AnalyteName'],
                                                    rapid_ws2['CS'],
                                                    rapid_ws2['ClinSigComment'],
                                                    rapid_ws2['RecordDate']))
                                            elif sapid[row_ws1]['AEONGO'] == "否" and ENcheck == False:
                                                if rapid_ws2['AnalyteName'] in AN_left:
                                                    AN_left.remove(rapid_ws2['AnalyteName'])
                                                    End_lines.append(rapid_ws2)
                                            if len(AN_left) == 0:
                                                ENcheck = True
                                                if len(AnalyteNamelist) == 1:
                                                    msg_e = ('Info:该AE记录结束日期与Lab页面第'
                                                    '{0}行{1} {2} {3} 匹配成功'.format(
                                                        row_ws2,
                                                        rapid_ws2['visitname'],
                                                        rapid_ws2['AnalyteName'],
                                                        rapid_ws2['CS']))
                                                else:
                                                    rsg = ','.join("{0} {1} {2}".format(i['visitname'], i['AnalyteName'], i['CS']) for i in End_lines)
                                                    msg_e = ("Info:该AE记录结束日期与Lab页面 {} 匹配成功".format(rsg))
                                
                                if not STcheck:
                                    msg_s = '\n'.join([msg_s, 'Error:该AE记录开始日期'
                                    '与第{0}行 {1} {2} {3} {4} {5}匹配失败'.format(
                                        row_ws2, rapid_ws2['visitname'],
                                        rapid_ws2['AnalyteName'], rapid_ws2['CS'],
                                        rapid_ws2['ClinSigComment'],
                                        rapid_ws2['RecordDate'])])
                                # if GR and not GRcheck:
                                #     msg = '\n'.join([msg, 'Error:该AE记录级别变化日期匹配失败'])
                                if EN and not ENcheck:
                                    msg_e = '\n'.join([msg_e, 'Error:该AE记录结束日期'
                                    '与第{0}行 {1} {2} {3} {4} 匹配失败'.format(
                                        row_ws2,rapid_ws2['visitname'],
                                        rapid_ws2['AnalyteName'], rapid_ws2['CS'],
                                        rapid_ws2['RecordDate'])])
                                mark(ws1, 'A', row_ws1, msg_s)
                                mark(ws1, 'B', row_ws1, msg_e)

                        if not AnalyteNamecheck:        
                            msg = 'Error:该患者在Lab页面无{}对应检查项'.format(AE_PT)
                            mark(ws1, 'A', row_ws1, msg)
                            continue

def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file                         

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--aesheet", default=r'AE|不良事件', help="Please set sheet name of ae")
    parser.add_argument("--labsheet", default=r'LAB|实验室检测', help="Please set sheet name of cb")

    files = get_files()
    file_name = get_a_file(files, "不良事件")
    ae_path = os.path.join(SHEETS_PATH, file_name)

    file_name = get_a_file(files, "LAB")
    lab_path = os.path.join(SHEETS_PATH, file_name)

    args = parser.parse_args()

    ae_sheet = args.aesheet
    lab_sheet = args.labsheet

    ae_pathlist = ae_path.split('.xlsx')
    lab_pathlist = lab_path.split('.xlsx')

    wb1savepath = ''.join([''.join([ae_pathlist[0], '_checkout']), '.xlsx'])
    wb2savepath = ''.join([''.join([lab_pathlist[0], '_checkout']), '.xlsx'])
    try:
        wb1 = openpyxl.load_workbook(ae_path)
        ws1 = wb1[ae_sheet]

        wb2 = openpyxl.load_workbook(lab_path)
        ws2 = wb2[lab_sheet]
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)

        data_ws1 = data1(ws1, keys1)
        data_ws2 = data2(ws2, keys2)

        ws2 = lab2ae(data_ws1, data_ws2, ws2)
        ws1 = ae2lab(data_ws1, data_ws2, ws1)

        wb1.save(wb1savepath)
        wb2.save(wb2savepath)
        

    finally:
        wb1.close()
        wb2.close()
