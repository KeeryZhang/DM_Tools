#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message

keys1list = [r'{change}','[Subject]', '[MHFLOC]', '[MHFCAT]', '[MHFDEGR]', '[MHFCST]', '[MHFTST]', '[MHFNST]', '[MHFMST]', '[MHFTYN]',
            '[MHFLOC1]', '[MHFLOC2]', '[MHFLOC3]', '[MHFLOC4]', '[MHFLOC5]', '[MHFLOC6]', '[MHFLOC7]', '[MHFLOC8]', '[MHFLOC9]', '[MHFLOC10]',
            '[MHFLOC11]', '[MHFLOC12]', '[MHFLOC13]', '[MHFLOC14]', '[MHFLOC15]', '[MHFLOC16]', '[MHFLOC17]', '[MHFLOC18]', '[MHFLOC19]',
            '[MHFLOC20]', '[MHFLOC21]', '[MHFLOC22]', '[MHFLOC23]', '[MHFLOC24]', '[MHFLOC25]', '[MHFLOC26]', '[MHFLOTH]', '[MHSFYN]']

keys2list = ['[MHSLOC]', '[MHSCAT]', '[MHSDEGR]', '[MHSCST]', '[MHSTST]', '[MHSNST]', '[MHSMST]', '[MHSTYN]',
            '[MHSLOC1]', '[MHSLOC2]', '[MHSLOC3]', '[MHSLOC4]', '[MHSLOC5]', '[MHSLOC6]', '[MHSLOC7]', '[MHSLOC8]', '[MHSLOC9]', '[MHSLOC10]',
            '[MHSLOC11]', '[MHSLOC12]', '[MHSLOC13]', '[MHSLOC14]', '[MHSLOC15]', '[MHSLOC16]', '[MHSLOC17]', '[MHSLOC18]', '[MHSLOC19]', 
            '[MHSLOC20]', '[MHSLOC21]', '[MHSLOC22]', '[MHSLOC23]', '[MHSLOC24]', '[MHSLOC25]', '[MHSLOC26]', '[MHSLOTH]']

SHEETS_PATH = "..\sheets"


def data(ws, keys1, keys2):
    data_ws1 = {}
    data_ws2 = {}
    for row in range(2, ws.max_row+1):
        if r'{change}' in keys1 and ws[keys1[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys1['[Subject]']+str(row)].value == None:
            continue
        
        keys1_tmp = deepcopy(keys1)
        keys2_tmp = deepcopy(keys2)

        subject = ws[keys1['[Subject]']+str(row)].value
        data_ws1.setdefault(subject, {})
        data_ws2.setdefault(subject, {})
        
        data_ws1[subject].setdefault('row', row)

        keys1_tmp.pop('[Subject]')
        keys1_tmp.pop(r'{change}')

        for key in keys1_tmp.keys():
            key_new = ''.join(filter(str.isalnum, key))
            value = ws[keys1_tmp[key]+str(row)].value
            if value == None:
                value = 0
            data_ws1[subject].setdefault(key_new, value)
        
        for key in keys2_tmp.keys():
            key_new = ''.join(filter(str.isalnum, key))
            value = ws[keys2_tmp[key]+str(row)].value
            if value == None:
                value = 0
            data_ws2[subject].setdefault(key_new, value)

    return data_ws1, data_ws2


def blxzdcheck(data_ws1, data_ws2, ws):
    ws.insert_cols(1)
    ws['A1'].value = '首次诊断一致性核查'

    for id in data_ws1:
        pid1 = data_ws1[id]
        pid2 = data_ws2[id]
        rsg = ''
        msg = ''
        checkpass = True
        if pid1['MHSFYN'] != '是':
            rsg = 'Info: 该行不进行首次诊断一致性核查'
        else:
            for key in data_ws2[id].keys():
                keylist = list(key)
                keylist[2] = 'F'
                kt = ''.join(keylist)
                if data_ws1[id][kt] == data_ws2[id][key]:
                    continue
                else:
                    checkpass = False
                    rsg += '{0}为{1}，{2}为{3} '.format(kt, data_ws1[id][kt], key, data_ws2[id][key])
            if checkpass:
                rsg = 'Info: 首次诊断一致性核查通过'
            else:
                msg = 'Error: 首次诊断一致性核查未通过。'
        msg = message(msg, rsg)
        mark(ws, "A", data_ws1[id]['row'], msg)


def yczycheck(data_ws1, data_ws2, ws):
    ws.insert_cols(1)
    ws['A1'].value = '远处转移一致性核查'

    for id in data_ws1:
        pid1 = data_ws1[id]
        pid2 = data_ws2[id]
        rsg = ''
        msg = ''
        checkpass = True

        sample = 'MHSLOC{}'
        newkeylist = list()
        for i in range(1,27):
            newkeylist.append(sample.format(i))

        for key in newkeylist:
            keylist = list(key)
            keylist[2] = 'F'
            kt = ''.join(keylist)
            if data_ws1[id][kt] == 1 and data_ws1[id][kt] != data_ws2[id][key]:
                checkpass = False
                rsg += '{0}为{1}，{2}为{3} '.format(kt, data_ws1[id][kt], key, data_ws2[id][key])

        if checkpass:
            rsg = 'Info: 远处转移一致性核查通过'
        else:
            msg = 'Error: 远处转移一致性核查未通过。'
        msg = message(msg, rsg)
        mark(ws, "A", data_ws1[id]['row'], msg)

def visitcheck(data_ws1, data_ws2, ws1):
    ws1.insert_cols(1)
    ws1['A6'].value = 'visit->研究药物给药核查结果'

    for id in data_ws1:
        pid_ws1 = data_ws1[id]
        if id not in data_ws2:
            iderror = True
        else:
            pid_ws2 = data_ws2[id]
            iderror = False

        for instance in pid_ws1:
            ipid_ws1 = pid_ws1[instance]
            msg = ''
            if 'C' not in instance:
                visitpass = True
            else:
                visitpass = False

            instanceerror = False

            if not visitpass and not iderror:    
                if instance not in pid_ws2:                    
                    instanceerror = True
                else:
                    ipid_ws2 = pid_ws2[instance]
                    instanceerror = False

            if not visitpass and not iderror and not instanceerror:
                if ipid_ws2['EXYN'] == "否" and ipid_ws1['drug'] == set():
                    exynpass = True
                else:
                    exynpass = False

            if visitpass:
                rsg = 'Info: 该行访视为 {} 无需匹配'.format(instance)
            elif iderror:
                rsg = 'Error: 该受试者{}信息在研究药物给药页面不存在'.format(id)
            elif instanceerror and ipid_ws1['skip'] == 'Y' and ipid_ws1['drug'] == set():
                rsg = 'Info: 该行跳过访视，访视 {} 在研究药物给药页面不存在'.format(instance)
            elif not instanceerror and ipid_ws1['skip'] == 'Y' and ipid_ws1['drug'] == set() and ipid_ws2['drug'] == set():
                rsg = 'Info: 该行跳过访视，药物记录在两边均为空'
            elif instanceerror:
                rsg = 'Error: 该受试者{}的访视{}信息在研究药物给药页面不存在'.format(id, instance)
            elif exynpass:
                rsg = 'Info: 该行未给药且使用药物为空'
            else:
                if ipid_ws1['drug'] == ipid_ws2['drug']:
                    rsg = 'Info: 该行使用药物匹配成功'
                else:
                    ws1diffws2 = ipid_ws1['drug'].difference(ipid_ws2['drug'])
                    ws2diffws1 = ipid_ws2['drug'].difference(ipid_ws1['drug'])
                    rsg = 'Error: 该行使用药物匹配失败，受试者 {0} 访视 {1}'.format(id, instance)            
                    if len(ws1diffws2) != 0:
                        rsg += ' visit页面多出 {}'.format(' '.join(str(x) for x in ws1diffws2))
                    if len(ws2diffws1) != 0:
                        rsg += ' 给药页面多出 {}'.format(' '.join(str(x) for x in ws2diffws1))

            msg = message(msg, rsg)
            mark(ws1, 'A', ipid_ws1['row'], msg)


def gycheck(data_ws1, data_ws2, ws1):
    ws1.insert_cols(1)
    ws1['A6'].value = '研究药物给药->visit核查结果'

    for id in data_ws1:
        pid_ws1 = data_ws1[id]
        if id not in data_ws2:
            iderror = True
        else:
            pid_ws2 = data_ws2[id]
            iderror = False

        for instance in pid_ws1:
            ipid_ws1 = pid_ws1[instance]
            msg = ''
            if 'C' not in instance:
                visitpass = True
            elif 'CC' in instance:
                visitpass = True
            else:
                visitpass = False

            if not visitpass and not iderror:
                if instance not in pid_ws2:
                    instanceerror = True
                else:
                    ipid_ws2 = pid_ws2[instance]
                    instanceerror = False

            if not visitpass and not iderror and not instanceerror:
                if ipid_ws1['EXYN'] == "否" and ipid_ws2['drug'] == set():
                    exynpass = True
                else:
                    exynpass = False

            if visitpass:
                rsg = 'Info: 该行访视为 {} 无需匹配'.format(instance)
            elif iderror:
                rsg = 'Error: 该受试者{}信息在受试者页面不存在'.format(id)
            elif ipid_ws1['EXYN'] == "否" and instanceerror:
                rsg = 'Info: 该行未给药且受试者页面不存在访视 {}'.format(instance)
            elif instanceerror:
                rsg = 'Error: 该受试者{}的访视{}信息在受试者页面不存在'.format(id, instance)
            elif exynpass:
                rsg = 'Info: 该行未给药且使用药物为空'            
            else:
                if ipid_ws1['drug'] == ipid_ws2['drug']:
                    rsg = 'Info: 该行使用药物匹配成功'
                else:
                    rsg = 'Error: 该行药物匹配失败，已在IRT受试者页面记录'
                    # ws1diffws2 = ipid_ws1['drug'].difference(ipid_ws2['drug'])
                    # ws2diffws1 = ipid_ws2['drug'].difference(ipid_ws1['drug'])
                    # rsg = 'Error: 该行使用药物匹配失败，受试者 {0} 访视 {1}'.format(id, instance)            
                    # if len(ws1diffws2) != 0:
                    #     rsg += ' 给药页面多出 {}'.format(' '.join(str(x) for x in ws1diffws2))
                    # if len(ws2diffws1) != 0:
                    #     rsg += ' visit页面多出 {}'.format(' '.join(str(x) for x in ws2diffws1))

            msg = message(msg, rsg)
            mark(ws1, 'A', ipid_ws1['row'], msg)


def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    # parser.add_argument("--visit", default=r'Copy of Subject_visit_report_KN046-3011629948672870.xlsx', help="Please add AE file name")
    # parser.add_argument("--gy", default=r'KN046-301_研究药物给药.xlsx', help="Please set sheet name of ae")
    parser.add_argument("--blxzd", default=r'MHDIAG|鳞状非小细胞肺癌病理学诊断', help="Please set sheet name of cb")
    parser.add_argument("--ywgy", default=r'EX|研究药物给药', help="Please set sheet name of cb")

    files = get_files()
    zd = get_a_file(files, "鳞状非小细胞肺癌病理学诊断")
    args = parser.parse_args()

    zd_path = os.path.join(SHEETS_PATH, zd)

    blxzd_sheet = args.blxzd

    zd_pathlist = zd_path.split('.xlsx')

    wbsavepath = ''.join([''.join([zd_pathlist[0], '_checkout']), '.xlsx'])

    try:
        wb = openpyxl.load_workbook(zd_path)
        ws = wb[blxzd_sheet]
                       
        keys1 = findkeyscolumn(ws, keys1list)
        keys2 = findkeyscolumn(ws, keys2list)

        data_ws1, data_ws2 = data(ws, keys1, keys2)
        
        blxzdcheck(data_ws1, data_ws2, ws)
        yczycheck(data_ws1, data_ws2, ws)

        wb.save(wbsavepath)

    finally:
        wb.close()