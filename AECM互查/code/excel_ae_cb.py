#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse


M2m = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
aerecordlist = ['[AEDSL1]','[AEDSL2]','[AEDSL3]','[AEDSL4]','[AEDSL5]']
keys1list = ['[Subject]','[AETERM]','[AESTDAT_RAW]','[AECONTRT]']
keys2list = ['[Subject]','[CMINDC]','[CMSTDAT_RAW]']


def parse_dmy(s):
    day_s,mon_s,year_s=s.split(' ')
    if day_s == 'UN':
        day_s = '1'
    if mon_s == 'UNK':
        mon_s = 'JAN'
    if year_s == '0000':
        year_s = '1970'

    return datetime(int(year_s),int(M2m[mon_s]),int(day_s))


def data1(ws,keys):
    data_ws = {}
    classify = {}
    for row in range(2,ws.max_row+1):
        treated = ws[keys[3]+str(row)].value

        if treated == '是' or treated == '否' :
            treatedset = set()
            msgset = set()
            patientId = ws[keys[0]+str(row)].value
            aeterm = ws[keys[1]+str(row)].value
            starttime = ws[keys[2]+str(row)].value
            st = parse_dmy(starttime)            
            msg = ""
            data_ws.setdefault(patientId,{})
            data_ws[patientId].setdefault(row,{'aeterm':aeterm, 'st':st, 'treated':treated, 'msg':msg})

            classify.setdefault(patientId,{})
            classify[patientId].setdefault(aeterm+str(st),[treatedset, msgset])
            classify[patientId][aeterm+str(st)][0].add(treated)
    return data_ws, classify


def data2(ws,keys,aerecord):
    data_ws = {}
    classify = {}
    for row in range(2,ws.max_row+1):
        cmindc = ws[keys[1]+str(row)].value

        if cmindc == '不良事件，请具体说明':
            patientId = ws[keys[0]+str(row)].value                
            starttime = ws[keys[2]+str(row)].value
            st = parse_dmy(starttime)
            data_ws.setdefault(patientId,{})
            data_ws[patientId].setdefault(row,{'st':st,'ae':{}})

            classify.setdefault(patientId,{})
            
            for ae in aerecord:
                aedsl = ws[ae+str(row)].value
                if aedsl == None:
                    continue
                else:
                    aedsl_list = aedsl.split(' - ')

                    aename = aedsl_list[-3]
                    aestarttime = aedsl_list[-2]                
                    ast = parse_dmy(aestarttime)
                    data_ws[patientId][row]['ae'].setdefault('ae'+str(aerecord.index(ae)+1),[aename, ast])
                    comb = aename+str(ast)
                    rowcheck = classify[patientId].setdefault(comb, [row])
                    if rowcheck[0] != row:
                        classify[patientId][comb].append(row)
    return data_ws, classify


def mark(ws, row, col, msg):
    ws[col+str(row)] = msg
    return


def crosscheck(data_ws1, classify1, classify2, ws1):
    ws1.insert_cols(1)
    ws1['A1'].value="检查结果"
    for id in data_ws1:
        pid = data_ws1[id]
        for row_ws1 in pid:
            pr1 = pid[row_ws1]
            if pr1['treated'] == '是':
                comb = pr1['aeterm']+str(pr1['st'])
                if id in classify2:
                    patient = classify2[id]
                    if comb in patient:
                        pr1['msg'] = 'Info:该不良事件匹配成功'
                        classify1[id][comb][1].add(pr1['msg'])
                        mark(ws1, row_ws1, 'A', pr1['msg'])
                    else:
                        pr1['msg'] = 'Error:该不良事件在合并用药中不存在'
                        mark(ws1,row_ws1,'A',pr1['msg'])  
                else:
                    pr1['msg'] = 'Error:该原因在合并用药中不存在'
                    mark(ws1, row_ws1, 'A', pr1['msg'])  
            else:
                continue

        for row_ws1 in pid:
            pr1 = pid[row_ws1]
            if pr1['treated'] == '否':
                comb = pr1['aeterm']+str(pr1['st'])
                if '是' in classify1[id][comb][0]:
                    if 'Info:该不良事件匹配成功' in classify1[id][comb][1]:
                        pr1['msg'] = 'Info:该不良事件存在合并治疗，且合并治疗匹配成功'
                        mark(ws1, row_ws1, 'A', pr1['msg'])
                    else:
                        pr1['msg'] = 'Error:该不良事件存在合并治疗，但合并治疗匹配失败'
                else:
                    if id in classify2:
                        patient = classify2[id]
                        if comb in patient:
                            pr1['msg'] = 'Error:该不良事件异常出现在合并用药，相关异常行数： '+' '.join(str(x) for x in patient[comb])
                            mark(ws1, row_ws1, 'A', pr1['msg'])
                        else:
                            pr1['msg'] = 'Info:该不良事件无合并治疗记录'
                            mark(ws1, row_ws1, 'A', pr1['msg'])
                    else:
                        pr1['msg'] = 'Info:该患者无合并用药记录'
                        mark(ws1, row_ws1, 'A', pr1['msg'])
            else:
                continue
    return ws1
    

def aetimecheck(data_ws2, ws2):
    if ws2['A1'].value != 'YD Comments':
        ws2.insert_cols(1)
        ws2['A1'].value = 'YD Comments'
    for pid in data_ws2:
        patient = data_ws2[pid]
        for row_ws2 in patient:
            processing = False
            pr = patient[row_ws2]
            for ae_ws2 in pr['ae']:
                pra = pr['ae'][ae_ws2]
                if pra[1] <= pr['st']:
                    mark(ws2, row_ws2, 'A', 'Pass')
                    processing = True
                    break
            if processing == False:
                mark(ws2, row_ws2, 'A', 'Error:合并用药开始日期早于所有AE开始日期')
    return ws2


def findkeyscolumn(ws, keyslist):
    keys = []
    for column in range(1, ws.max_column+1):
        row_letter = get_column_letter(column)
        for key in keyslist:
            if key in ws[row_letter+'1'].value:
                keys.append(row_letter)
    return keys


if __name__=="__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ae", default=r'E:\python\DM_Tool\AECM互查\CIBI308E301_AE_YD20210727.xlsx', help="Please add AE file full path")
    parser.add_argument("--aesheet", default=r'AE001_1|不良事件-不包括输液反应及免疫相关不良事件', help="Please set sheet name of ae")
    parser.add_argument("--cb", default=r'E:\python\DM_Tool\AECM互查\CIBI308E301_合并用药_YD20210727.xlsx', help="Please add CB file full path")
    parser.add_argument("--cbsheet", default=r'CM001_4|既往及合并药物治疗', help="Please set sheet name of cb")
    parser.add_argument("--flow", default="all", help="Please state the flow you need to run")

    args = parser.parse_args()

    ae_path = args.ae
    ae_sheet = args.aesheet
    cb_path = args.cb
    cb_sheet = args.cbsheet
    flow = args.flow

    ae_pathlist = ae_path.split('.')
    cb_pathlist = cb_path.split('.')

    wb1savepath = '.'.join([''.join([ae_pathlist[0], '_checkout']), ae_pathlist[1]])
    wb2savepath = '.'.join([''.join([cb_pathlist[0], '_checkout']), cb_pathlist[1]])
    try:
        wb1 = openpyxl.load_workbook(ae_path)
        ws1 = wb1.get_sheet_by_name(ae_sheet)

        wb2 = openpyxl.load_workbook(cb_path)
        ws2 = wb2.get_sheet_by_name(cb_sheet)
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)
        aerecord = findkeyscolumn(ws2, aerecordlist)

        data_ws1,classify1 = data1(ws1, keys1)
        data_ws2,classify2 = data2(ws2, keys2, aerecord)

        if flow == 'crosscheck':
            ws1 = crosscheck(data_ws1, classify1, classify2, ws1)
            wb1.save(wb1savepath)
        elif flow == 'aetimecheck':
            ws2 = aetimecheck(data_ws2, ws2)
            wb2.save(wb2savepath)
        elif flow == 'all':
            ws1 = crosscheck(data_ws1, classify1, classify2, ws1)
            ws2 = aetimecheck(data_ws2, ws2)
            wb1.save(wb1savepath)
            wb2.save(wb2savepath)

    finally:
        wb1.close()
        wb2.close()