#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy
import threading

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message, parse_dmy

keyslist = ["[RecordDate]", "[PEDAT]", "[VSDAT]", "[ECOGDAT]", "[UCGDAT]", "[EGDAT]", "[LBDAT]", "[LBDAT]", "[LBDAT]", "[URDAT]", "[UPRODAT]", "[RPBDAT]", "[RPUDAT]", "[VIRODAT]", "[EXSTDAT]", "[EXSTDAT]", "[EXSTDAT]", "[EXSTDAT]", "[PCCDAT]", "[PCDAT]", "[ADADAT]", "[ADACDAT]", "[TLDAT]", "[NTLDAT]", "[NWTDAT]", "[IMEDAT]", "[RSDAT]","[AESTDAT]", "[AECSTDAT]", "[CMSTDAT]", "[PRSTDAT]", "[FRSTDAT]", "[FSDAT]", "[FRSTDAT]", "[FCSTDAT]", "[FOSTDAT]", "[TCIDAT]", "[DSSDDAT]", "[FUDAT]", "[DSDAT]", "[LBDAT]", "[BNPDAT]"]

SHEETS_PATH = "..\sheets"

def data(ws, keys, checkdate):
    max_column = ws.max_column + 1

    ws[get_column_letter(max_column)+"1"].value = 'Date Check Result ({})'.format(checkdate.strftime('%Y-%m-%d'))
    for row in range(2,ws.max_row+1):        
        row_check = False
        earlier = False
        for key in keys:
            date = ws[keys[key]+str(row)].value
            if date is not None:
                row_check = True
                if date <= checkdate:
                    earlier = True
                    if key == "[RecordDate]":
                        break
        
        if row_check:
            if earlier:
                mark(ws, get_column_letter(max_column), row, "Y")
            else:
                mark(ws, get_column_letter(max_column), row, "N")
    return

def get_files(path):
    files_raw = os.listdir(path)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file

def date_convert(raw_date):
    yy, mm, dd = raw_date.split("-")
    return datetime(int(yy), int(mm), int(dd))

if __name__=="__main__":
    parser = argparse.ArgumentParser()
    
    # parser.add_argument("--date", default=r'', 
    #                     help="Please set sheet name of ae")
    # parser.add_argument("--cmsheet", default=r'CM|既往及合并用药', 
    #                     help="Please set sheet name of cm")
    # parser.add_argument("--flow", default="all", 
    #                     help="Please state the flow you need to run")

    files = get_files(SHEETS_PATH)

    file_name = get_a_file(files, "_datadumplistings_")
    ddl_path = os.path.join(SHEETS_PATH, file_name)
    checkdate = date_convert(input("Please provide check date (YYYY-MM-DD): "))
    # args = parser.parse_args()

    ddl_pathlist = ddl_path.split('.xlsx')
    wbsavepath = ''.join([''.join([ddl_pathlist[0], '_checkout']), '.xlsx'])

    try:
        wb = openpyxl.load_workbook(ddl_path)
        sheetslist = wb.sheetnames

        threads = []
        for sheet in sheetslist:
            ws = wb[sheet]
            keys = findkeyscolumn(ws, keyslist)
            t = threading.Thread(target=data, args=(ws, keys, checkdate))
            t.start()
            threads.append(t)

        for t in threads:
            t.join()

        wb.save(wbsavepath)
    finally:
        wb.close()