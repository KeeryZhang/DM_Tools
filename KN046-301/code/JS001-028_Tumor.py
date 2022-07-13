#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message

SHEETS_PATH = "..\sheets"

keys1list = ['[Subject]', '[InstanceName]', '[TMAYN]', '[TMASPD]', '[TMADAT]']
keys2list = ['[Subject]', '[InstanceName]', '[IOTAPERF]', '[IOTADAT]', '[IOTATL]']


def data(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        tmp_keys = deepcopy(keys)
        # if ws[tmp_keys[r'{change}']+str(row)].value == 'deleted':
        #     continue
        if ws[tmp_keys['[Subject]']+str(row)].value == None:
            continue
                
        Subject = ws[tmp_keys['[Subject]']+str(row)].value
        InstanceName = ws[tmp_keys['[InstanceName]']+str(row)].value
        
        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(InstanceName, {})
        data_ws[Subject][InstanceName].setdefault(row, {})

        # tmp_keys.pop(r'{change}')
        tmp_keys.pop('[Subject]')
        tmp_keys.pop('[InstanceName]')

        for key in tmp_keys:
            data_ws[Subject][InstanceName][row].update({key:ws[tmp_keys[key]+str(row)].value})

    return data_ws


def bbzpretriage(data_ws, data_ws2, ws, TU):
    id_delete = set()
    for id in data_ws:
        pid = data_ws[id]
        instance_delete = set()
        for instance in pid:
            ipid = pid[instance]
            row_delete = []
            for row in ipid:
                msg = ''
                rsg = ''
                ripid = ipid[row]
                if exist(id, data_ws2):
                    pid_ws2 = data_ws2[id]
                    if exist(instance, pid_ws2):
                        ipid_ws2 = pid_ws2[instance]
                        if TU == '靶病灶' and ripid['[TMAYN]'] == '否':
                            for row_ws4 in ipid_ws2:
                                ripid_ws2 = ipid_ws2[row_ws4]
                            if ripid_ws2['[IOTAPERF]'] == '否' or 'NA' in ripid_ws2['[IOTATL]']:
                                rsg = 'Info:该行{}评估为否，在Recist页面存在IOTAPERF为否或IOTATL为NA'.format(TU)                                
                            else:
                                rsg = 'Error:该受试者 {0} 在访视 {1} 中{2}评估为否，但在Recist页面IOTAPERF不为否或IOTATL不为NA'.format(id, instance, TU)
                            row_delete.append(row)
                        elif TU == '靶病灶' and ripid['[TMAYN]'] == None:
                            rsg = 'Error:该行{}评估为空'.format(TU)                            
                            row_delete.append(row)
                        elif TU == '靶病灶' and ripid['[TMASPD]'] == None:
                            rsg = 'Error:该行靶病灶直径和为空'
                            row_delete.append(row)
                    else:
                        if '筛选期（-28天）' not in instance:
                            rsg = 'Error:该受试者在Recist页面无访视 {} 信息'.format(instance)
                            instance_delete.add(instance)
                else:
                    if '筛选期（-28天）' not in instance:
                        rsg = 'Error:该受试者在Recist页面不存在'                       
                    else:
                        rsg = 'Info:该受试者处于筛选期在Recist页面不存在'
                    id_delete.add(id)
                msg = message(msg, rsg)
                mark(ws, 'A', row, msg)

            if len(row_delete) > 0:
                for row in row_delete:
                    ipid.pop(row)

            if len(ipid) == 0:
                instance_delete.add(instance)
        
        if len(instance_delete) > 0:
            for instance in instance_delete:
                pid.pop(instance)

        if len(pid) == 0:
            id_delete.add(id)
    
    if len(id_delete) > 0:
        for id in id_delete:
            data_ws.pop(id)

    return data_ws


def pid_revert(pid):
    pid_normal = {}
    pid_cc = {}
    for instance in pid:
        rows = []
        rows_cc = []
        if 'CC' in instance or 'cc' in instance:
            pid_cc.setdefault(instance, {})
            pid_cc[instance].setdefault('TMADAT_error', False)
            for row in pid[instance]:
                if pid[instance][row]['[TMADAT]'] is None:
                    pid_cc[instance]['TMADAT_error'] = True
                rows_cc.append(row)
                for key in pid[instance][row]:
                    pid_cc[instance].setdefault(key, pid[instance][row][key])
            pid_cc[instance].setdefault('rows',rows_cc)
        else:
            pid_normal.setdefault(instance, {})
            pid_normal[instance].setdefault('TMADAT_error', False)
            for row in pid[instance]:
                if pid[instance][row]['[TMADAT]'] is None:
                    pid_normal[instance]['TMADAT_error'] = True
                rows.append(row)
                for key in pid[instance][row]:
                    pid_normal[instance].setdefault(key, pid[instance][row][key])
            pid_normal[instance].setdefault('rows',rows)

    return pid_normal, pid_cc


def bbzresult(InstanceName, crossbase, crosschecklist, crossinstancelist):
    index = crossinstancelist.index(InstanceName)
    check = crosschecklist[index]
    checklist = crosschecklist[:index+1]
    checklist.append(crossbase)
    TMASPD_min = min(checklist)
    zerocheck = True
    if TMASPD_min != 0:
        PRcheck = (check - crossbase)/crossbase
        PDcheck = (check - TMASPD_min)/TMASPD_min
    else:
        zerocheck = False
        
    if zerocheck:    
        if PDcheck >= 0.2 and abs(check - TMASPD_min) >= 5:
            result = 'PD'
        elif abs(PRcheck) >= 0.3:
            result = 'PR'
        else:
            result = 'SD'
    else:
        result = '0'
    return result

    
def bbzpidresult(pid_sorted, pid_ws2):
    crossbase = int()
    crosschecklist = list()
    crossinstancelist = list()
    for i in range(0, len(pid_sorted)):
        instance = pid_sorted[i][0]
        p_ws1 = pid_sorted[i][1]
        TMASPD = p_ws1['[TMASPD]']
        if '筛选期（-28天）' in instance:
            crossbase = TMASPD
        else:
            crosschecklist.append(TMASPD)
            crossinstancelist.append(instance)

    for i in range(0, len(pid_sorted)):
        instance = pid_sorted[i][0]
        p_ws1 = pid_sorted[i][1]
        TMASPD = p_ws1['[TMASPD]']
        msg = ''
        rsg = ''
        if '筛选期（-28天）' in instance:
            crossbase = TMASPD
            rsg = 'Info:该行为筛选期，跳过比较'
        else:
            result = bbzresult(instance, crossbase, crosschecklist, crossinstancelist)
            for row_ws4 in pid_ws2[instance]:
                if result in pid_ws2[instance][row_ws4]['[IOTATL]']:
                    rsg = 'Info:该行靶病灶结果匹配成功'
                elif result == '0':
                    rsg = 'Warn:此次检测数值为零，需提供说明'
                else:
                    rsg = 'Error:该行靶病灶结果应为 {}，与Recist页面第 {} 行匹配失败'.format(result, row_ws4)
        msg = message(msg, rsg)
        for row in p_ws1['rows']:
            mark(ws1, 'A', row, msg)


def bbzpidcheck(pid_ws1_ori, pid_ws2_ori, ws1):
    pid_ws1 = deepcopy(pid_ws1_ori)
    pid_ws2 = deepcopy(pid_ws2_ori)
    
    pid_normal, pid_cc = pid_revert(pid_ws1)

    if pid_normal != {}:
        pid_normal = error_check(pid_normal, ws1)
        pid_normal = sorted(pid_normal.items(), key = lambda time:time[1]['[TMADAT]'])
        bbzpidresult(pid_normal, pid_ws2)
            
    if pid_cc != {}:
        pid_cc = error_check(pid_cc, ws1)
        pid_cc = sorted(pid_cc.items(), key = lambda time:time[1]['[TMADAT]'])
        bbzpidresult(pid_cc, pid_ws2)   
    return 


def error_check(pid, ws):
    ''' Remove lines whose [TMADAT] is empty, and mark error '''
    pid_copy = deepcopy(pid)
    for instance in pid:
        if pid[instance]['TMADAT_error']:
            for row in pid[instance]['rows']:
                msg = "Error: 该访视存在 TMADAT 缺失"
                mark(ws, 'A', row, msg)
            pid_copy.pop(instance)
    return pid_copy


def bbzcheck(data_ws1_ori, data_ws2, ws1):
    ws1.insert_cols(1)
    ws1['A1'].value = '靶病灶检查结果'

    data_ws1 = deepcopy(data_ws1_ori)  

    data_ws1 = bbzpretriage(data_ws1, data_ws2, ws1, '靶病灶')

    for id in data_ws1:
        pid_ws1 = data_ws1[id]
        pid_ws2 = data_ws2[id]
        bbzpidcheck(pid_ws1, pid_ws2, ws1)

    return


def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file


if __name__ == "__main__":
    files = get_files()
    parser = argparse.ArgumentParser()
    parser.add_argument("--bbz", default=r'TMA001_2|肿瘤评估-靶病灶', help="Please set sheet name of ae")
    parser.add_argument("--recist", default=r'IOTA001_1|实体瘤疗效评估（RECIST V1.1）', help="Please set sheet name of cb")

    args = parser.parse_args()

    cancer = get_a_file(files, "JS001-028-III-SCLC_肿瘤评估")
    cancer_path = os.path.join(SHEETS_PATH, cancer)
    bbz_sheet = args.bbz
    recist_sheet = args.recist

    cancer_pathlist = cancer_path.split('.xlsx')

    wbsavepath = ''.join([''.join([cancer_pathlist[0], '_checkout']), '.xlsx'])
    try:
        wb = openpyxl.load_workbook(cancer_path)
        ws1 = wb[bbz_sheet]
        ws2 = wb[recist_sheet]
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)

        data_ws1 = data(ws1, keys1)
        data_ws2 = data(ws2, keys2)
        
        bbzcheck(data_ws1, data_ws2, ws1)

        wb.save(wbsavepath)

    finally:
        wb.close()
