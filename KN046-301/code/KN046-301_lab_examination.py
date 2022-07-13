#!/usr/bin/python
# -*- coding:UTF-8 -*-

from multiprocessing.sharedctypes import Value
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from datetime import datetime
import os
import sys
import argparse
import copy
from copy import deepcopy
import pdb

sys.path.append("..\..")
from tool_lib.outs.Codelist_map import Codelist_map as Codelist_map
from tool_lib.outs.Coderevert_map import Coderevert_map_withsymbol as Coderevert_map

from tool_lib.utils import findkeyscolumn, parse_dmy, mark


keyslist = [r'{change}', '[Subject]', '[InstanceName]', '[RecordDate]', '[AnalyteName]',
            '[NumericValue]', '[LabFlag]', '[ClinSigValue]']


SHEETS_PATH = "..\sheets"


class Data:
    def __init__(self, item=None):
        self.item = item
        self.data = dict()

    def add_item(self, item_key):
        pass

    def find_item_by_id(self, item_key):
        if item_key not in self.data:
            self.add_item(item_key)
        return self.data[item_key]


class Data_sheet(Data):
    def add_item(self, subject):
        self.data.setdefault(subject, self.Data_subject(subject))

    class Data_subject(Data):
        def add_item(self, analytename):
            self.data.setdefault(analytename, self.Data_analyte(analytename))
        
        @property
        def subject(self):
            return self.item

        class Data_analyte(Data):
            def __init__(self, item=None):
                super().__init__(item)
                self.base = None

            def add_item(self, row, **kwargs):
                self.data.setdefault(row, self.Data_row(row, kwargs))

            @property
            def analytename(self):
                return self.item

            def get_base(self):
                if self.base is None:
                    for data_row in self.data.values():
                        if data_row.instance == "筛选期" and data_row.sigvalue == "异常有临床意义":
                            self.base = data_row
                            break
                        elif data_row.instance == "筛选期" and data_row.sigvalue != "异常有临床意义":
                            self.base = "No base"
                            break
                return self.base

            class Data_row:
                """ Store data of each row """
                def __init__(self, row, data):
                    self.row = row
                    self.data = data
                
                @property
                def instance(self):
                    return self.data["instance"]
                @property
                def date(self):
                    return self.data["date"]
                
                @property
                def value(self):
                    return self.data["value"]

                @property
                def flag(self):
                    return self.data["flag"]

                @property
                def sigvalue(self):
                    return self.data["sigvalue"]


def data(ws, keys):
    data_ws = Data_sheet()
    for row in range(2, ws.max_row+1):
        if ws[keys[r'{change}']+str(row)] == 'deleted' or \
           ws[keys['[Subject]']+str(row)] == None:
            continue

        subject = ws[keys['[Subject]']+str(row)].value
        analytename = ws[keys['[AnalyteName]']+str(row)].value

        data_subject = data_ws.find_item_by_id(subject)
        data_analyte = data_subject.find_item_by_id(analytename)
        data_analyte.add_item(row,
                              instance=ws[keys['[InstanceName]']+str(row)].value,
                              date=ws[keys['[RecordDate]']+str(row)].value,
                              value=ws[keys['[NumericValue]']+str(row)].value,
                              flag=ws[keys['[LabFlag]']+str(row)].value,
                              sigvalue=ws[keys['[ClinSigValue]']+str(row)].value)

    return data_ws


def lab_examination_check(data_ws, ws):
    ws.insert_cols(1)
    ws['A1'].value = 'DM Comment'

    for data_subject in data_ws.data.values():
        for data_analyte in data_subject.data.values():
            base = data_analyte.get_base()
            if base == "No base":
                continue

            for data_row in data_analyte.data.values():
                msg = ""
                if data_row.sigvalue == "异常无临床意义" and data_row.date >= base.date:
                    if base.flag == "+":
                        if data_row.value >= base.value :
                            msg = "Error: 该行 %s 检测值 %s 大于筛选期 %s 且不为 异常有临床意义" % (data_analyte.analytename, data_row.value, base.value)
                    elif base.flag == "-":
                        if data_row.value <= base.value:
                            msg = "Error: 该行 %s 检测值 %s 小于筛选期 %s 且不为 异常有临床意义" % (data_analyte.analytename, data_row.value, base.value)
                    mark(ws, "A", data_row.row, msg)

    return ws


def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files


def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file                         


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--labsheet", default=r'LAB|实验室检测', help="Please set sheet name of cb")

    files = get_files()
    file_name = get_a_file(files, "实验室检测")
    lab_path = os.path.join(SHEETS_PATH, file_name)

    args = parser.parse_args()

    lab_sheet = args.labsheet

    lab_pathlist = lab_path.split('.xlsx')

    wbsavepath = ''.join([''.join([lab_pathlist[0], '_checkout']), '.xlsx'])
    try:
        wb = openpyxl.load_workbook(lab_path)
        ws = wb[lab_sheet]
               
        keys = findkeyscolumn(ws, keyslist)

        data_ws = data(ws, keys)

        ws1 = lab_examination_check(data_ws, ws)

        wb.save(wbsavepath)        

    finally:
        wb.close()
