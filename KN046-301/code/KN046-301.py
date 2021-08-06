#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message

keys1list = [r'{change}', '[Subject]', '[InstanceName]', '[TLYN]', '[TLDIAT]', '[TLDAT]', '[TLMETHOD]', '[TLLNKID]']
keys2list = [r'{change}', '[Subject]', '[InstanceName]', '[NTLYN]', '[NTLDAT]', '[NTLORRES]', '[NTLLNKID]', '[NTLMTHOD]']
keys3list = [r'{change}', '[Subject]', '[InstanceName]', '[NWTLEYN]']
keys4list = [r'{change}', '[Subject]', '[InstanceName]', '[RSYN]', '[RSDAT]', '[TRGRESP]', '[NTRGRESP]', '[NEWLIND]']


def data(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        tmp_keys = deepcopy(keys)
        if ws[tmp_keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[tmp_keys['[Subject]']+str(row)].value == None:
            continue
                
        Subject = ws[tmp_keys['[Subject]']+str(row)].value
        InstanceName = ws[tmp_keys['[InstanceName]']+str(row)].value
        
        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(InstanceName, {})
        data_ws[Subject][InstanceName].setdefault(row, {})

        tmp_keys.pop(r'{change}')
        tmp_keys.pop('[Subject]')
        tmp_keys.pop('[InstanceName]')

        for key in tmp_keys:
            data_ws[Subject][InstanceName][row].update({key:ws[tmp_keys[key]+str(row)].value})

        # data_ws[Subject][InstanceName][row].update({'messaged':False})
    return data_ws


def bbzpretriage(data_ws1, data_ws4, ws1):
    id_delete = set()
    for id in data_ws1:
        pid = data_ws1[id]
        instance_delete = set()
        for instance in pid:
            ipid = pid[instance]
            row_delete = []
            for row in ipid:
                msg = ''
                ripid = ipid[row]
                if exist(id, data_ws4):
                    pid_ws4 = data_ws4[id]
                    if exist(instance, pid_ws4):
                        ipid_ws4 = pid_ws4[instance]
                        if ripid['[TLYN]'] == '否':
                            for row_ws4 in ipid_ws4:
                                ripid_ws4 = ipid_ws4[row_ws4]
                            if ripid_ws4['[RSYN]'] == '否' or 'NA' in ripid_ws4['[TRGRESP]']:
                                rsg = 'Info:该行靶病灶评估为否，在Recist页面存在RSYN为否或TRGRESP为NA'
                                msg = message(msg, rsg)
                            else:
                                rsg = 'Error:该受试者 {0} 在访视 {1} 中靶病灶评估为否，但在Recist页面RSYN不为否或TRGRESP不为NA'.format(id, instance)
                                msg = message(msg, rsg)
                            row_delete.append(row)
                        elif ripid['[TLYN]'] == None:
                            rsg = 'Error:该行靶病灶评估为空'
                            msg = message(msg, rsg)
                            row_delete.append(row)
                    else:
                        if '筛选期' not in instance:
                            rsg = 'Error:该受试者在Recist页面无访视 {} 信息'.format(instance)
                            msg = message(msg, rsg)
                            instance_delete.add(instance)
                else:
                    rsg = 'Error:该受试者 {} 在Recist页面不存在'.format(id)
                    msg = message(msg, rsg)
                    id_delete.add(id)
                    
                mark(ws1, 'A', row, msg)        

            if len(row_delete) > 0:
                for row in row_delete:
                    ipid.pop(row)

            if len(ipid) == 0:
                instance_delete.add(instance)
        
        if len(instance_delete) > 0:
            for instance in instance_delete:
                pid.pop(instance)

        if len(pid) == 0:
            id_delete.add(id)
    
    if len(id_delete) > 0:
        for id in id_delete:
            data_ws1.pop(id)

    return data_ws1


def pid_revert(pid):
    pid_normal = {}
    pid_cc = {}
    for instance in pid:
        rows = []
        if 'CC' in instance or 'cc' in instance:
            pid_cc.setdefault(instance, {})
            
            for row in pid[instance]:
                rows.append(row)
                for key in pid[instance][row]:
                    pid_cc[instance].setdefault(key, pid[instance][row][key])
            pid_cc[instance].setdefault('rows',rows)

        else:
            pid_normal.setdefault(instance, {})
            for row in pid[instance]:
                rows.append(row)
                for key in pid[instance][row]:
                    pid_normal[instance].setdefault(key, pid[instance][row][key])
            pid_normal[instance].setdefault('rows',rows)

    return pid_normal, pid_cc


def bbzresult(InstanceName, crossbase, crosschecklist, crossinstancelist):
    index = crossinstancelist.index(InstanceName)
    check = crosschecklist[index]
    checklist = crosschecklist[:index+1]
    checklist.append(crossbase)
    TMASPD_min = min(checklist)
    zerocheck = True
    if TMASPD_min != 0:
        PRcheck = (check - crossbase)/crossbase
        PDcheck = (check - TMASPD_min)/TMASPD_min
    else:
        zerocheck = False
        
    if zerocheck:    
        if PDcheck >= 0.2 and abs(check - TMASPD_min) >= 5:
            result = 'PD'
        elif abs(PRcheck) >= 0.3:
            result = 'PR'
        else:
            result = 'SD'
    else:
        result = '0'
    return result

    
def bbzpidcheck(pid_ws1_ori, pid_ws4_ori, ws1):
    pid_ws1 = deepcopy(pid_ws1_ori)
    pid_ws4 = deepcopy(pid_ws4_ori)
    
    pid_normal, pid_cc = pid_revert(pid_ws1)

    if pid_normal != {}:
        pid_normal = sorted(pid_normal.items(), key = lambda time:time[1]['[TLDAT]'])

        for i in range(0, len(pid_normal)):
            instance = pid_normal[i][0]
            p_ws1 = pid_normal[i][1]
            msg = ''
            if '筛选期' in instance:
                rsg = 'Info:该行为筛选期，跳过比较'
                msg = message(msg, rsg)
            else:
                if 

            for row in p_ws1['rows']:
                mark(ws1, 'A', row, msg)

            ####################################################

    if pid_cc != {}:
        pid_cc = sorted(pid_cc.items(), key = lambda time:time[1]['[TLDAT]'])

        for i in range(0, len(pid_cc)):
            pass    
    return 

def bbzcheck(data_ws1_ori, data_ws4, ws1):
    ws1.insert_cols(1)
    ws1['A1'].value = '靶病灶检查结果'

    data_ws1 = deepcopy(data_ws1_ori)  

    data_ws1 = bbzpretriage(data_ws1, data_ws4, ws1)

    for id in data_ws1:
        pid_ws1 = data_ws1[id]
        pid_ws4 = data_ws4[id]
        bbzpidcheck(pid_ws1, pid_ws4, ws1)

    return





def fbbzcheck(data_ws2, data_ws4, ws2):

    return


def xbzcheck(data_ws3, data_ws4, ws3):

    return


def methodcheck(data_ws, ws):

    return

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--cancer", default=r'KN046-301_肿瘤评估_20210803.xlsx', help="Please add AE file name")
    parser.add_argument("--bbz", default=r'TUTL|肿瘤评价-靶病灶（RECIST 1.1）', help="Please set sheet name of ae")
    parser.add_argument("--fbbz", default=r'TUNTL|肿瘤评价-非靶病灶（RECIST 1.1）', help="Please set sheet name of cb")
    parser.add_argument("--xbz", default=r'TUNEWTL|肿瘤评价-新病灶（RECIST 1.1）', help="Please set sheet name of cb")
    parser.add_argument("--recist", default=r'RS|总体疗效评价（RECIST 1.1）', help="Please set sheet name of cb")

    args = parser.parse_args()

    cancer_path = os.path.join(r'..\sheets', args.cancer)    
    bbz_sheet = args.bbz
    fbbz_sheet = args.fbbz
    xbz_sheet = args.xbz
    recist_sheet = args.recist

    cancer_pathlist = cancer_path.split('.xlsx')

    wbsavepath = ''.join([''.join([cancer_pathlist[0], '_checkout']), '.xlsx'])
    try:
        wb = openpyxl.load_workbook(cancer_path)
        ws1 = wb[bbz_sheet]
        ws2 = wb[fbbz_sheet]
        ws3 = wb[xbz_sheet]
        ws4 = wb[recist_sheet]
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)
        keys3 = findkeyscolumn(ws3, keys3list)
        keys4 = findkeyscolumn(ws4, keys4list)

        data_ws1 = data(ws1, keys1)
        data_ws2 = data(ws2, keys2)
        data_ws3 = data(ws3, keys3)
        data_ws4 = data(ws4, keys4)
        
        # data_ws1 = data(ws1, keys1)

        bbzcheck(data_ws1, data_ws4, ws1)
        fbbzcheck(data_ws2, data_ws4, ws2)
        xbzcheck(data_ws3, data_ws4, ws3)

        methodcheck(data_ws1, ws1)
        methodcheck(data_ws2, ws2)

        wb.save(wbsavepath)

    finally:
        wb.close()