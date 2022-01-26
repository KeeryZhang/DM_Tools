#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message

SHEETS_PATH = "..\sheets"

keys1list = [r'{change}', '[Subject]', '[InstanceName]', '[TLYN]', '[TLDIAT]', '[TLDAT]', '[TLMETHOD]', '[TLLNKID]', '[TLRSND]']
keys2list = [r'{change}', '[Subject]', '[InstanceName]', '[NTLYN]', '[NTLDAT]', '[NTLORRES]', '[NTLLNKID]', '[NTLMTHOD]']
keys3list = [r'{change}', '[Subject]', '[InstanceName]', '[NWTLEYN]']
keys4list = [r'{change}', '[Subject]', '[InstanceName]', '[RSYN]', '[RSDAT]', '[TRGRESP]', '[NTRGRESP]', '[NEWLIND]']


def data(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        tmp_keys = deepcopy(keys)
        if ws[tmp_keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[tmp_keys['[Subject]']+str(row)].value == None:
            continue
                
        Subject = ws[tmp_keys['[Subject]']+str(row)].value
        InstanceName = ws[tmp_keys['[InstanceName]']+str(row)].value
        
        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(InstanceName, {})
        data_ws[Subject][InstanceName].setdefault(row, {})

        tmp_keys.pop(r'{change}')
        tmp_keys.pop('[Subject]')
        tmp_keys.pop('[InstanceName]')

        for key in tmp_keys:
            data_ws[Subject][InstanceName][row].update({key:ws[tmp_keys[key]+str(row)].value})

    return data_ws


def bbzpretriage(data_ws, data_ws4, ws, TU):
    
    id_delete = set()
    for id in data_ws:
        pid = data_ws[id]
        instance_delete = set()
        for instance in pid:
            ipid = pid[instance]
            row_delete = []
            for row in ipid:
                msg = ''
                rsg = ''
                ripid = ipid[row]
                if exist(id, data_ws4):
                    pid_ws4 = data_ws4[id]
                    if exist(instance, pid_ws4):
                        ipid_ws4 = pid_ws4[instance]
                        if TU == '靶病灶' and ripid['[TLYN]'] == '否':
                            for row_ws4 in ipid_ws4:
                                ripid_ws4 = ipid_ws4[row_ws4]
                            if ripid_ws4['[RSYN]'] == '否' or 'NA' in ripid_ws4['[TRGRESP]']:
                                rsg = 'Info:该行{}评估为否，在Recist页面存在RSYN为否或TRGRESP为NA'.format(TU)                                
                            else:
                                rsg = 'Error:该受试者 {0} 在访视 {1} 中{2}评估为否，但在Recist页面RSYN不为否或TRGRESP不为NA'.format(id, instance, TU)
                            row_delete.append(row)
                        elif TU == '靶病灶' and ripid['[TLYN]'] == None:
                            rsg = 'Error:该行{}评估为空'.format(TU)                            
                            row_delete.append(row)

                        if TU == '非靶病灶' and ripid['[NTLYN]'] == '否':
                            for row_ws4 in ipid_ws4:
                                ripid_ws4 = ipid_ws4[row_ws4]
                            if ripid_ws4['[NTRGRESP]'] == '基线无非靶病灶':
                                rsg = 'Info:该行{}评估为否，对应Recist页面中基线无非靶病灶'.format(TU)
                            else:
                                rsg = 'Error:该受试者 {0} 在访视 {1} 中{2}评估为否，但在Recist页面NTRGRESP不为基线无非靶病灶'.format(id, instance, TU)
                            row_delete.append(row)
                        elif TU == '非靶病灶' and ripid['[NTLYN]'] == None:
                            rsg = 'Error:该行{}评估为空'.format(TU)
                            row_delete.append(row)

                        if TU == '新病灶' and ripid['[NWTLEYN]'] == None:
                            rsg = 'Error:该行{}评估为空'.format(TU)
                            row_delete.append(row)
                    else:
                        if '筛选期' not in instance:
                            rsg = 'Error:该受试者在Recist页面无访视 {} 信息'.format(instance)
                            instance_delete.add(instance)
                        elif TU == '非靶病灶' and '筛选期' in instance:
                            rsg = 'Info:筛选期无Recist对应记录'
                            instance_delete.add(instance)
                else:
                    if '筛选期' not in instance:
                        rsg = 'Error:该受试者在Recist页面不存在'                       
                    else:
                        rsg = 'Info:该受试者处于筛选期在Recist页面不存在'
                    id_delete.add(id)
                msg = message(msg, rsg)
                mark(ws, 'A', row, msg)        

            if len(row_delete) > 0:
                for row in row_delete:
                    ipid.pop(row)

            if len(ipid) == 0:
                instance_delete.add(instance)
        
        if len(instance_delete) > 0:
            for instance in instance_delete:
                pid.pop(instance)

        if len(pid) == 0:
            id_delete.add(id)
    
    if len(id_delete) > 0:
        for id in id_delete:
            data_ws.pop(id)

    return data_ws


def pid_revert(pid):
    pid_normal = {}
    pid_cc = {}
    for instance in pid:
        rows = []
        rows_cc = []
        if 'CC' in instance or 'cc' in instance:
            pid_cc.setdefault(instance, {})
            pid_cc[instance].setdefault('TLDAT_error', False)
            for row in pid[instance]:
                if pid[instance][row]['[TLDAT]'] is None:
                    pid_cc[instance]['TLDAT_error'] = True
                rows_cc.append(row)
                for key in pid[instance][row]:
                    pid_cc[instance].setdefault(key, pid[instance][row][key])
            pid_cc[instance].setdefault('rows',rows_cc)
        else:
            pid_normal.setdefault(instance, {})
            pid_normal[instance].setdefault('TLDAT_error', False)
            for row in pid[instance]:
                if pid[instance][row]['[TLDAT]'] is None:
                    pid_normal[instance]['TLDAT_error'] = True
                rows.append(row)
                for key in pid[instance][row]:
                    pid_normal[instance].setdefault(key, pid[instance][row][key])
            pid_normal[instance].setdefault('rows',rows)

    return pid_normal, pid_cc


def bbzresult(InstanceName, crossbase, crosschecklist, crossinstancelist):
    index = crossinstancelist.index(InstanceName)
    check = crosschecklist[index]
    checklist = crosschecklist[:index+1]
    checklist.append(crossbase)
    TLDIAT_min = min(checklist)
    zerocheck = True
    if TLDIAT_min != 0:
        PRcheck = (check - crossbase)/crossbase
        PDcheck = (check - TLDIAT_min)/TLDIAT_min
    else:
        zerocheck = False
        
    if zerocheck:    
        if PDcheck >= 0.2 and abs(check - TLDIAT_min) >= 5:
            result = 'PD'
        elif abs(PRcheck) >= 0.3:
            result = 'PR'
        else:
            result = 'SD'
    else:
        result = '0'
    return result

    
def bbzpidresult(pid_sorted, pid_ws4):
    crossbase = int()
    crosschecklist = list()
    crossinstancelist = list()
    for i in range(0, len(pid_sorted)):
        instance = pid_sorted[i][0]
        p_ws1 = pid_sorted[i][1]
        TLDIAT = p_ws1['[TLDIAT]']
        if '筛选期' in instance:
            crossbase = TLDIAT                
        else:
            crosschecklist.append(TLDIAT)
            crossinstancelist.append(instance)

    for i in range(0, len(pid_sorted)):
        instance = pid_sorted[i][0]
        p_ws1 = pid_sorted[i][1]
        TLDIAT = p_ws1['[TLDIAT]']
        msg = ''
        rsg = ''
        if '筛选期' in instance:
            crossbase = TLDIAT
            rsg = 'Info:该行为筛选期，跳过比较'
        else:
            result = bbzresult(instance, crossbase, crosschecklist, crossinstancelist)
            for row_ws4 in pid_ws4[instance]:
                if result in pid_ws4[instance][row_ws4]['[TRGRESP]']:
                    rsg = 'Info:该行靶病灶结果匹配成功'
                elif p_ws1['[TLRSND]'] == "不明确的" and pid_ws4[instance][row_ws4]['[TRGRESP]']:
                    rsg = "Info:该受试者该访视靶病灶不可评估，匹配成功"
                elif result == '0':
                    rsg = 'Warn:此次检测数值为零，需提供说明'
                else:
                    rsg = 'Error:该行靶病灶结果应为 {}，与Recist页面第 {} 行匹配失败'.format(result, row_ws4)
        msg = message(msg, rsg)
        for row in p_ws1['rows']:
            mark(ws1, 'A', row, msg)


def bbzpidcheck(pid_ws1_ori, pid_ws4_ori, ws1):
    pid_ws1 = deepcopy(pid_ws1_ori)
    pid_ws4 = deepcopy(pid_ws4_ori)
    
    pid_normal, pid_cc = pid_revert(pid_ws1)

    if pid_normal != {}:
        pid_normal = error_check(pid_normal, ws1)
        pid_normal = sorted(pid_normal.items(), key = lambda time:time[1]['[TLDAT]'])
        bbzpidresult(pid_normal, pid_ws4)
            
    if pid_cc != {}:
        pid_cc = error_check(pid_cc, ws1)
        pid_cc = sorted(pid_cc.items(), key = lambda time:time[1]['[TLDAT]'])
        bbzpidresult(pid_cc, pid_ws4)   
    return 


def error_check(pid, ws):
    ''' Remove lines whose [TLDAT] is empty, and mark error '''
    pid_copy = deepcopy(pid)
    for instance in pid:
        if pid[instance]['TLDAT_error']:
            for row in pid[instance]['rows']:
                msg = "Error: 该访视存在 TLDAT 缺失"
                mark(ws, 'A', row, msg)
            pid_copy.pop(instance)
    return pid_copy


def bbzcheck(data_ws1_ori, data_ws4, ws1):
    ws1.insert_cols(1)
    ws1['A1'].value = '靶病灶检查结果'

    data_ws1 = deepcopy(data_ws1_ori)  

    data_ws1 = bbzpretriage(data_ws1, data_ws4, ws1, '靶病灶')

    for id in data_ws1:
        pid_ws1 = data_ws1[id]
        pid_ws4 = data_ws4[id]
        bbzpidcheck(pid_ws1, pid_ws4, ws1)

    return


def fbbzresult(ipid):
    NTLORRES_set = set()
    result = ''
    for row_ws2 in ipid:
        ripid_ws2 = ipid[row_ws2]
        NTLORRES_set.add(ripid_ws2['[NTLORRES]'])
    if '明确的进展' in NTLORRES_set:
        result = 'PD'
    elif '可见' in NTLORRES_set:
        result = r'非CR/非PD'
    elif '不能评估' in NTLORRES_set:
        result = 'NE'
    elif len(NTLORRES_set) > 0 and len(NTLORRES_set.union({'不可见'})) == 1:
        result = 'CR'
    elif len(NTLORRES_set) == 0:
        result = '0'
    return result


def fbbzpidcheck(pid_ws2_ori, pid_ws4_ori, ws2):
    pid_ws2 = deepcopy(pid_ws2_ori)
    pid_ws4 = deepcopy(pid_ws4_ori)
    
    for instance in pid_ws2:
        ipid_ws2 = pid_ws2[instance]
        ipid_ws4 = pid_ws4[instance]
        result = fbbzresult(ipid_ws2)
        msg = ''
        rsg = ''
        for row_ws4 in ipid_ws4:
            ripid_ws4 = ipid_ws4[row_ws4]
            if result in ripid_ws4['[NTRGRESP]']:
                rsg = 'Info:该行结果与非靶病灶匹配成功'
            elif result == '0':
                rsg = 'Error:该行非靶病灶评估为空'
            else:
                rsg = 'Error:该行非靶病灶结果应为 {} ，Recist结果为 {} ，与本行匹配失败'.format(result, ripid_ws4['[NTRGRESP]'])
        msg = message(msg, rsg)

        for row_ws2 in ipid_ws2:
            ripid_ws2 = ipid_ws2[row_ws2]
            if ripid_ws2['[NTLORRES]'] == None:
                msg = 'Error:该行非靶病灶评估为空'
            mark(ws2, 'A', row_ws2, msg)
    return


def fbbzcheck(data_ws2_ori, data_ws4, ws2):
    ws2.insert_cols(1)
    ws2['A1'].value = '非靶病灶检查结果'

    data_ws2 = deepcopy(data_ws2_ori)  

    data_ws2 = bbzpretriage(data_ws2, data_ws4, ws2, '非靶病灶')

    for id in data_ws2:
        pid_ws2 = data_ws2[id]
        pid_ws4 = data_ws4[id]
        fbbzpidcheck(pid_ws2, pid_ws4, ws2)
    return


def xbzpidcheck(pid_ws3, pid_ws4, ws3):
    for instance in pid_ws3:
        ipid_ws3 = pid_ws3[instance]
        ipid_ws4 = pid_ws4[instance]
        for row_ws3 in ipid_ws3:
            msg = ''
            rsg = ''
            ismatch = False
            ripid_ws3 = ipid_ws3[row_ws3]
            for row_ws4 in ipid_ws4:
                ripid_ws4 = ipid_ws4[row_ws4]
            if ripid_ws3['[NWTLEYN]'] == '是':
                if ripid_ws4['[NEWLIND]'] == '有':
                    ismatch = True
            elif ripid_ws3['[NWTLEYN]'] == '否':
                if ripid_ws4['[NEWLIND]'] == '无':
                    ismatch = True
            
            if ismatch:
                rsg = 'Info:该行结果与新病灶匹配成功'
            else:
                rsg = 'Error:该行新病灶结果应为 {} ，Recist结果为 {} ，与本行匹配失败'.format(ripid_ws3['[NWTLEYN]'], ripid_ws4['[NEWLIND]'])
            msg = message(msg, rsg)
            mark(ws3, 'A', row_ws3, msg)
    return


def xbzcheck(data_ws3_ori, data_ws4, ws3):
    ws3.insert_cols(1)
    ws3['A1'].value = '新病灶检查结果'

    data_ws3 = deepcopy(data_ws3_ori)  

    data_ws3 = bbzpretriage(data_ws3, data_ws4, ws3, '新病灶')

    for id in data_ws3:
        pid_ws3 = data_ws3[id]
        pid_ws4 = data_ws4[id]
        xbzpidcheck(pid_ws3, pid_ws4, ws3)
    return


def methodpretriage(data_ws, ws, TU):
    if TU == '靶病灶':
        idname = '[TLLNKID]'
        method = '[TLMETHOD]'
        ifcheck = '[TLYN]'
    else:
        idname = '[NTLLNKID]'
        method = '[NTLMTHOD]'
        ifcheck = '[NTLYN]'

    data_ws_normal_revert = {}
    data_ws_cc_revert = {}
    for id in data_ws:
        pid = data_ws[id]
        for instance in pid:
            ipid = pid[instance]
            row_delete = []
            for row in ipid:
                msg = ''
                rsg = ''
                ripid = ipid[row]
                if ripid[ifcheck] == '否':
                    rsg = 'Info:该行不进行{}评估'.format(TU)
                    row_delete.append(row)
                elif ripid[idname] == None and ripid[method] == None:
                    rsg = 'Error:该行无{}编号且检查方法为空'.format(TU)
                    row_delete.append(row)
                elif ripid[idname] == None:
                    rsg = 'Error:该行无{}编号'.format(TU)
                    row_delete.append(row)
                elif ripid[method] == None:
                    rsg = 'Error:该行{}检查方法为空'.format(TU)
                    row_delete.append(row)                
                else:
                    if 'CC' in instance or 'cc' in instance:
                        data_ws_cc_revert.setdefault(id, {})
                        data_ws_cc_revert[id].setdefault(ripid[idname], [])
                        if '筛选期' in instance:
                            data_ws_cc_revert[id][ripid[idname]].insert(0, (row, ripid[method], instance))
                        else:
                            data_ws_cc_revert[id][ripid[idname]].append((row, ripid[method], instance))
                    else:
                        data_ws_normal_revert.setdefault(id, {})
                        data_ws_normal_revert[id].setdefault(ripid[idname], [])
                        if '筛选期' in instance:
                            data_ws_normal_revert[id][ripid[idname]].insert(0, (row, ripid[method], instance))
                        else:
                            data_ws_normal_revert[id][ripid[idname]].append((row, ripid[method], instance))
                msg = message(msg, rsg)
                mark(ws, 'A', row, msg)
                     

    return data_ws_normal_revert, data_ws_cc_revert


def methodprocess(data_ws, ws, TU):
    for id in data_ws:
        pid = data_ws[id]
        for idname in pid:
            ipid = pid[idname]
            hasbase = False
            if '筛选期' in ipid[0][2]:
                checkbase = ipid[0][1]
                for i in range(0, len(ipid)):
                    msg = ''
                    if '筛选期' in ipid[i][2]:
                        rsg = 'Info:该行为受试者病灶编号{}筛选期，作为对比基准'.format(idname)
                    else:
                        if ipid[i][1] == checkbase:
                            rsg = 'Info:该行检测方法与筛选期一致'
                        else:
                            rsg = 'Error:该行检测方法为 {}，筛选期检测方法为 {}，匹配失败'.format(ipid[i][1], checkbase)
                    msg = message(msg, rsg)
                    mark(ws, 'A', ipid[i][0], msg)
            else:
                msg = ''
                rsg = 'Error:该受试者病灶编号{}无筛选期检查，无法进行对比'.format(idname)
                msg = message(msg, rsg)
                for i in range(0, len(ipid)):
                    mark(ws, 'A', ipid[i][0], msg)
    return

def methodcheck(data_ws_ori, ws, TU):
    ws.insert_cols(1)
    ws['A1'].value = '检查方法对比结果'

    data_ws = deepcopy(data_ws_ori)

    data_normal, data_cc = methodpretriage(data_ws, ws, TU)

    methodprocess(data_normal, ws, TU)
    methodprocess(data_cc, ws, TU)
    return

def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file


if __name__ == "__main__":
    files = get_files()
    parser = argparse.ArgumentParser()
    # parser.add_argument("--cancer", default=r'KN046-301_肿瘤评估_20210826.xlsx', help="Please add AE file name")
    parser.add_argument("--bbz", default=r'TUTL|肿瘤评价-靶病灶（RECIST 1.1）', help="Please set sheet name of ae")
    parser.add_argument("--fbbz", default=r'TUNTL|肿瘤评价-非靶病灶（RECIST 1.1）', help="Please set sheet name of cb")
    parser.add_argument("--xbz", default=r'TUNEWTL|肿瘤评价-新病灶（RECIST 1.1）', help="Please set sheet name of cb")
    parser.add_argument("--recist", default=r'RS|总体疗效评价（RECIST 1.1）', help="Please set sheet name of cb")

    args = parser.parse_args()

    # cancer_path = os.path.join(r'..\sheets', args.cancer)
    cancer = get_a_file(files, "KN046-301_肿瘤评估")
    cancer_path = os.path.join(SHEETS_PATH, cancer)
    bbz_sheet = args.bbz
    fbbz_sheet = args.fbbz
    xbz_sheet = args.xbz
    recist_sheet = args.recist

    cancer_pathlist = cancer_path.split('.xlsx')

    wbsavepath = ''.join([''.join([cancer_pathlist[0], '_checkout']), '.xlsx'])
    try:
        wb = openpyxl.load_workbook(cancer_path)
        ws1 = wb[bbz_sheet]
        ws2 = wb[fbbz_sheet]
        ws3 = wb[xbz_sheet]
        ws4 = wb[recist_sheet]
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)
        keys3 = findkeyscolumn(ws3, keys3list)
        keys4 = findkeyscolumn(ws4, keys4list)

        data_ws1 = data(ws1, keys1)
        data_ws2 = data(ws2, keys2)
        data_ws3 = data(ws3, keys3)
        data_ws4 = data(ws4, keys4)
        
        bbzcheck(data_ws1, data_ws4, ws1)
        fbbzcheck(data_ws2, data_ws4, ws2)
        xbzcheck(data_ws3, data_ws4, ws3)

        methodcheck(data_ws1, ws1, '靶病灶')
        methodcheck(data_ws2, ws2, '非靶病灶')

        wb.save(wbsavepath)

    finally:
        wb.close()