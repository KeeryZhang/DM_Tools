#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message, parse_dmy

keys1list = [r'{change}','[Subject]', '[RecordPosition]','[CMTRT]', '[CMSTDAT_RAW]', '[CMENDAT_RAW]','[CMROUT]']

SHEETS_PATH = "..\sheets"


def data(ws, keys1):
    data_ws1 = {}

    for row in range(2, ws.max_row+1):
        if r'{change}' in keys1 and ws[keys1[r'{change}']+str(row)].value == 'deleted':
            continue

        if ws[keys1['[Subject]']+str(row)].value == None:
            continue

        subject = ws[keys1['[Subject]']+str(row)].value
        log = ws[keys1['[RecordPosition]']+str(row)].value
        cmname = ws[keys1['[CMTRT]']+str(row)].value
        start = ws[keys1['[CMSTDAT_RAW]']+str(row)].value
        startdate = parse_dmy(start, "/")
        end = ws[keys1['[CMENDAT_RAW]']+str(row)].value
        enddate = parse_dmy(end, "/")
        cmrout = ws[keys1['[CMROUT]']+str(row)].value

        data_ws1.setdefault(subject, {})        
        data_ws1[subject].setdefault(cmname,{})
        data_ws1[subject][cmname].setdefault(cmrout,{})
        data_ws1[subject][cmname][cmrout].setdefault(row, {"log": log, "startdate" :startdate,"enddate": enddate})
    return data_ws1

def cmrepeatcheck(data_ws1, ws):
    ws.insert_cols(1)
    ws['A1'].value = '药物名称查重'

    for subject in data_ws1:
        pid = data_ws1[subject]
        for cmname in pid:
            cpid = pid[cmname]
            for cmrout in cpid:
                opid = cpid[cmrout]
                pid_sorted = sorted(opid.items(), key = lambda time:time[1]['startdate'])
                # if len(pid_sorted) == 1:
                #     mark(ws, "A", pid_sorted[0][0], "Info: 该药物 {0} 只有一行记录".format(cmname))
                for i in range(len(pid_sorted)-1):
                    pid_startmin = pid_sorted[i]
                    msg = ''
                    for pid_check in pid_sorted[i+1:]:
                        rsg = ''
                        if pid_check[1]['startdate'] < pid_startmin[1]['enddate']:
                            rsg = "#{0} 和 #{1} 的药物 {2} 日期有重叠，请核实是否重复记录，谢谢。".format(pid_check[1]['log'], pid_startmin[1]['log'], cmname)
                        msg = message(msg, rsg)
                    mark(ws, "A", pid_startmin[0], msg)
                    # if len(pid_sorted) > 1:
                    #     mark(ws, "A", pid_sorted[-1][0], "")

def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--cm", default=r'CM|既往及合并用药', help="Please set sheet name of cb")

    files = get_files()
    cmrepeat = get_a_file(files, "合并用药repeat")
    args = parser.parse_args()

    cmrepeat_path = os.path.join(SHEETS_PATH, cmrepeat)

    cmrepeat_sheet = args.cm

    cmrepeat_pathlist = cmrepeat_path.split('.xlsx')

    wbsavepath = ''.join([''.join([cmrepeat_pathlist[0], '_checkout']), '.xlsx'])

    try:
        wb = openpyxl.load_workbook(cmrepeat_path)
        ws = wb[cmrepeat_sheet]
                       
        keys1 = findkeyscolumn(ws, keys1list)
        data_ws1 = data(ws, keys1)
        cmrepeatcheck(data_ws1, ws)

        wb.save(wbsavepath)
    finally:
        wb.close()