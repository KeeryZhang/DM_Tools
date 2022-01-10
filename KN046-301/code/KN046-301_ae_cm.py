#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message, parse_dmy


keys1list = [r'{change}', '[Subject]', '[AETERM]', '[AESTDAT_RAW]', '[AENTRT]', '[RecordPosition]', '[AEENDAT_RAW]']
keys2list = [r'{change}', '[Subject]', '[CMRSNAE]', '[CMSTDAT_RAW]', '[CMENDAT_RAW]']
aerecordlist = ['[CMAENO1]', '[CMAENO2]', '[CMAENO3]', '[CMAENO4]', '[CMAENO5]']

SHEETS_PATH = "..\sheets"

def data1(ws, keys):
    data_ws = {}
    for row in range(2,ws.max_row+1):
        if r'{change}' in keys and ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue

        if ws[keys['[Subject]']+str(row)].value == None:
            continue

        subject = ws[keys['[Subject]']+str(row)].value
        AETERM = ws[keys['[AETERM]']+str(row)].value
        AESTDAT_RAW = ws[keys['[AESTDAT_RAW]']+str(row)].value
        AESTDAT = parse_dmy(AESTDAT_RAW, '/')
        AEENDAT_RAW = ws[keys['[AEENDAT_RAW]']+str(row)].value
        AEENDAT = parse_dmy(AEENDAT_RAW, '/')
        AENTRT = ws[keys['[AENTRT]']+str(row)].value
        RecordPosition = ws[keys['[RecordPosition]']+str(row)].value

        data_ws.setdefault(subject, {})
        data_ws[subject].setdefault(row, {'AETERM': AETERM,
                                          'AESTDAT': AESTDAT,
                                          'AEENDAT': AEENDAT,
                                          'AENTRT': AENTRT,
                                          'RecordPosition': RecordPosition})
    return data_ws

def data2(ws, keys, aerecord):
    data_ws = {}
    for row in range(2,ws.max_row+1):
        if r'{change}' in keys and ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue

        if ws[keys['[Subject]']+str(row)].value == None:
            continue

        subject = ws[keys['[Subject]']+str(row)].value
        CMRSNAE = ws[keys['[CMRSNAE]']+str(row)].value
        CMSTDAT_RAW = ws[keys['[CMSTDAT_RAW]']+str(row)].value
        CMENDAT_RAW = ws[keys['[CMENDAT_RAW]']+str(row)].value
        CMSTDAT = parse_dmy(CMSTDAT_RAW, '/')
        CMENDAT = parse_dmy(CMENDAT_RAW, '/')
        data_ws.setdefault(subject, {})
        data_ws[subject].setdefault(row, {'CMRSNAE': CMRSNAE,
                                          'CMSTDAT': CMSTDAT,
                                          'CMENDAT': CMENDAT,
                                          'CMAENO': {}})

        for ae in aerecordlist:
            CMAENO = ws[aerecord[ae]+str(row)].value

            if CMAENO is None:
                data_ws[subject][row]['CMAENO'][ae] = None
            else:
                aelist = CMAENO.split(' - ')
                aename = aelist[-2]
                aetime_raw = aelist[-1]
                aetime = parse_dmy(aetime_raw, '/')
                aelog = aelist[0]
                data_ws[subject][row]['CMAENO'].setdefault(ae, {'aename': aename,
                                                                'aetime': aetime,
                                                                'aelog': aelog})
    return data_ws

def aecheck(data_ws1, data_ws2, ws1):
    """ aecheck """
    ws1.insert_cols(1)
    ws1['A1'].value = 'AE->CM'
    for id in data_ws1:
        pid1 = data_ws1[id]

        for row_ws1 in pid1:
            rpid1 = pid1[row_ws1]
            msg = ''
            ae_check = False
            time_check = False
            exist_check = False
            
            if id not in data_ws2:
                iderror = True
            else:
                pid2 = data_ws2[id]
                iderror = False
            
            if not iderror and rpid1['AENTRT'] == "是":
                for row_ws2 in pid2:
                    rpid2 = pid2[row_ws2]
                    CM_all = rpid2['CMAENO']

                    for CMAENO in CM_all:
                        if CM_all[CMAENO] is None:
                            continue

                        if rpid1['AETERM'] == CM_all[CMAENO]['aename']:
                            ae_check = True

                            if rpid1['AESTDAT'] == CM_all[CMAENO]['aetime']:
                                time_check = True
                                break

                    if time_check == True:
                        break

            elif not iderror and rpid1['AENTRT'] == "否":
                exist_check = True

                for row_ws2 in pid2:
                    rpid2 = pid2[row_ws2]
                    CM_all = rpid2['CMAENO']

                    for CMAENO in CM_all:
                        if CM_all[CMAENO] is None:
                            continue

                        if rpid1['AETERM'] == CM_all[CMAENO]['aename']:
                            ae_check = True
                            if rpid1['AESTDAT'] == CM_all[CMAENO]['aetime']:
                                time_check = True
                                break

                    if time_check == True:
                        break                            

            if rpid1['AENTRT'] == "是" and iderror:
                msg = "Error: 该受试者 {0} 信息在CM页面不存在".format(id)
            elif rpid1['AENTRT'] == "否" and iderror:
                msg = "Info: 该受试者 {0} 无需药物治疗且在CM页面不存在"
            elif exist_check and ae_check and time_check:
                msg = "Error: 该受试者 {0} 不良事件 {1} 无需药物治疗，在CM页面第{2}行匹配异常".format(id, rpid1['AETERM'], row_ws2)
            elif exist_check:
                msg = "Info: 该受试者 {0} 不良事件 {1} 无需药物治疗，且CM页面无匹配".format(id, rpid1['AETERM'])
            elif not ae_check or not time_check:
                msg = "Error: 该受试者 {0} 不良事件 {1} 在CM页面匹配失败".format(id, rpid1['AETERM'])
            elif ae_check and time_check:
                msg = "Info: 该受试者 {0} 不良事件 {1} 匹配成功".format(id, rpid1['AETERM'])

            mark(ws1, "A", row_ws1, msg)

def cmcheck(data_ws1, data_ws2, ws2):
    """ cmcheck """
    ws2.insert_cols(1)
    ws2['A1'].value = 'CM->AE'
    for id in data_ws2:
        pid2 = data_ws2[id]

        for row_ws2 in pid2:
            rpid2 = pid2[row_ws2]
            CM_all = rpid2['CMAENO']
            rsg = ''
            msg = ''
            iderror = False
            
            empty_number = 0

            if rpid2['CMRSNAE'] == 1 and id not in data_ws1:
                iderror = True
                msg = "Error: 该受试者 {0} 信息在AE页面不存在".format(id)
            elif rpid2['CMRSNAE'] == 0 and id not in data_ws1:
                iderror = True
                msg = "Info: 该受试者 {0} 无不良事件且在AE页面不存在".format(id)
            else:
                pid1 = data_ws1[id]
            
            if not iderror and rpid2['CMRSNAE'] == 1:
                for CMAENO in CM_all:
                    if CM_all[CMAENO] is None:
                        continue
                    else:
                        match_error = True
                        for row_ws1 in pid1:
                            rpid1 = pid1[row_ws1]
                            if CM_all[CMAENO]['aename'] == rpid1['AETERM'] and \
                                CM_all[CMAENO]['aetime'] == rpid1['AESTDAT']:
                                if rpid1['AENTRT'] == "是":
                                    match_error = False
                                    break
                                else:
                                    match_error = False
                                    rsg = "Error: 用药原因为“不良事件# {0} {1}”, 但是该不良事件是否进行药物治疗选择“否”，请核实，谢谢。" \
                                        .format(CM_all[CMAENO]['aelog'], CM_all[CMAENO]['aename'])
                                    break
                        if match_error:
                            rsg = "Error: 该受试者 {0} 不良事件 {1} 在AE页面匹配失败".format(id, CM_all[CMAENO]['aename'])

                    msg = message(msg, rsg)
                if msg == '':
                    msg = "Info: 该受试者 {0} 所有不良事件均匹配成功".format(id)
            elif not iderror and rpid2['CMRSNAE'] == 0:
                for CMAENO in CM_all:
                    if CM_all[CMAENO] is None:
                        empty_number += 1
                        continue
                    else:
                        rsg = "Error: 本行不应有不良事件，但{0}内容不为空".format(CMAENO)
                        break
                if empty_number == 5:
                    msg = "Info: 本行无不良事件，无需比较"
                else:
                    msg = message(msg, rsg)
            
            mark(ws2, "A", row_ws2, msg)

def timecheck(data_ws2, ws2):
    """ cm time check """
    ws2.insert_cols(1)
    ws2['A1'].value = 'CM time check'
    for id in data_ws2:
        pid2 = data_ws2[id]
        for row_ws2 in pid2:
            rpid2 = pid2[row_ws2]
            time_check = False
            cm_time = rpid2['CMSTDAT']
            empty_number = 0
            empty_error = False
            exist_error = False
            
            if rpid2['CMRSNAE'] == 1:
                CM_all = rpid2['CMAENO']

                for CMAENO in CM_all:
                    if CM_all[CMAENO] is None:
                        empty_number += 1
                        continue

                    if CM_all[CMAENO]['aetime'] <= cm_time:
                        time_check = True
                        break

                if empty_number == 5:
                    empty_error = True
            else:
                CM_all = rpid2['CMAENO']

                for CMAENO in CM_all:
                    if CM_all[CMAENO] is not None:
                        exist_error = True
                        break
            
            if time_check:
                msg = "Info: 本行存在AE日期不晚于CM日期"
            elif empty_error:
                msg = "Error: 本行应有不良事件，但内容为空"
            elif exist_error:
                msg = "Error: 本行不应有不良事件，但{0}内容不为空".format(CMAENO)
            elif rpid2['CMRSNAE'] == 0:
                msg = "Info: 本行无不良事件，无需比较"
            else:
                rsg = ""
                tmp = ""
                for CMAENO in CM_all:
                    if CM_all[CMAENO] is None:
                        continue
                    tmp = ' '.join([CM_all[CMAENO]['aelog'], CM_all[CMAENO]['aename']])
                    if rsg == "":
                        split = ''
                    else:
                        split = '，'
                    rsg = split.join([rsg, tmp])
                msg = "Error: 用药原因为“不良事件# {0}”, 但是该不良事件开始日期晚于用药开始日期，请核实，谢谢。(请同时核查MH和预防给药)" \
                                        .format(rsg)

            mark(ws2, "A", row_ws2, msg)

def endtimecheck(data_ws1, data_ws2, ws2):
    """ AE end time check """
    ws2.insert_cols(1)
    ws2['A1'].value = 'AE end time check'
    for id in data_ws2:
        pid2 = data_ws2[id]
        for row_ws2 in pid2:
            rpid2 = pid2[row_ws2]
            time_check = False
            time_check_num = 0
            # cm_time = rpid2['CMENDAT']
            empty_number = 0
            empty_error = False
            exist_error = False
            
            if rpid2['CMRSNAE'] == 1:
                CM_all = rpid2['CMAENO']

                for CMAENO in CM_all:
                    if CM_all[CMAENO] is None:
                        empty_number += 1
                        continue

                    # if CM_all[CMAENO]['aetime'] <= cm_time:
                    #     time_check = True
                    #     break
                    pid1 = data_ws1[id]
                    for row in pid1:
                        rpid1 = pid1[row]
                        if CM_all[CMAENO]['aelog'] == rpid1['RecordPosition']:
                            if rpid2['CMSTDAT'] > rpid1['AEENDAT']:
                                time_check_num += 1
                            break

                if empty_number == 5:
                    empty_error = True
                if time_check_num < len(CM_all):
                    time_check = True
            else:
                CM_all = rpid2['CMAENO']

                for CMAENO in CM_all:
                    if CM_all[CMAENO] is not None:
                        exist_error = True
                        break
            
            if time_check:
                msg = "Info: 本行存在AE结束日期早于CM开始日期"
            elif empty_error:
                msg = "Error: 本行应有不良事件，但内容为空"
            elif exist_error:
                msg = "Error: 本行不应有不良事件，但{0}内容不为空".format(CMAENO)
            elif rpid2['CMRSNAE'] == 0:
                msg = "Info: 本行无不良事件，无需比较"
            else:
                rsg = ""
                tmp = ""
                for CMAENO in CM_all:
                    if CM_all[CMAENO] is None:
                        continue
                    tmp = ' '.join([CM_all[CMAENO]['aelog'], CM_all[CMAENO]['aename']])
                    if rsg == "":
                        split = ''
                    else:
                        split = '，'
                    rsg = split.join([rsg, tmp])
                msg = "Error: 用药原因为“不良事件# {0}”，但是该不良事件结束日期{1}早于用药开始日期，请核实，谢谢。" \
                    .format(rsg, 'AEENDAT')

            mark(ws2, "A", row_ws2, msg)

def get_files(path):
    files_raw = os.listdir(path)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file
        

if __name__=="__main__":
    parser = argparse.ArgumentParser()
    
    parser.add_argument("--aesheet", default=r'AE|不良事件', 
                        help="Please set sheet name of ae")
    parser.add_argument("--cmsheet", default=r'CM|既往及合并用药', 
                        help="Please set sheet name of cm")
    parser.add_argument("--flow", default="all", 
                        help="Please state the flow you need to run")

    files = get_files(SHEETS_PATH)

    file_name = get_a_file(files, "_合并用药_")
    cm_path = os.path.join(SHEETS_PATH, file_name)

    args = parser.parse_args()

    ae_sheet = args.aesheet
    cm_sheet = args.cmsheet
    flow = args.flow

    cm_pathlist = cm_path.split('.xlsx')
    wb2savepath = ''.join([''.join([cm_pathlist[0], '_checkout']), '.xlsx'])

    try:
        wb2 = openpyxl.load_workbook(cm_path)
        ws1 = wb2[ae_sheet]
        ws2 = wb2[cm_sheet]

        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)
        aerecord = findkeyscolumn(ws2, aerecordlist)

        data_ws1 = data1(ws1, keys1)
        data_ws2 = data2(ws2, keys2, aerecord)

        aecheck(data_ws1, data_ws2, ws1)
        cmcheck(data_ws1, data_ws2, ws2)
        timecheck(data_ws2, ws2)
        endtimecheck(data_ws1, data_ws2, ws2)
        wb2.save(wb2savepath)

    finally:
        wb2.close()