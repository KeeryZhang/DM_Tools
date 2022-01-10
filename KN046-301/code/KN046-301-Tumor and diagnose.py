#!/usr/bin/python
# -*- coding:UTF-8 -*-


import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message

keys1list = [r'{change}','[Subject]','[MHSLOC]','[MHSLOC2]', '[MHSLOC3]', '[MHSLOC4]', '[MHSLOC5]', '[MHSLOC7]', '[MHSLOC9]',
            '[MHSLOC12]', '[MHSLOC13]', '[MHSLOC15]', '[MHSLOC17]', '[MHSLOC19]', '[MHSLOC20]', '[MHSLOC22]']
keys2list = [r'{change}','[Subject]','[InstanceName]','[TLLOC]']
keys3list = [r'{change}','[Subject]','[InstanceName]','[NTLLOC]']

LLOC2SLOC = {'肺': '[MHSLOC]', 'CNS(脑/脊柱/眼)':'[MHSLOC2]','骨':'[MHSLOC3]','肝脏':'[MHSLOC4]','肾上腺':'[MHSLOC5]',
             '皮肤/软组织':'[MHSLOC7]','结直肠':'[MHSLOC9]','小肠':'[MHSLOC9]','胃':'[MHSLOC12]','支气管淋巴结':'[MHSLOC13]',
             '隆突下淋巴结转移':'[MHSLOC13]','肺门淋巴结':'[MHSLOC13]','纵膈淋巴结':'[MHSLOC13]','前斜角肌淋巴结':'[MHSLOC13]',
             '锁骨上区淋巴结':'[MHSLOC13]','局部/区域/分段淋巴结':'[MHSLOC13]','远处转移淋巴结':'[MHSLOC13]','乳腺':'[MHSLOC15]',
             '胸膜/胸腔渗出液':'[MHSLOC17]','膀胱':'[MHSLOC19]','头颈部(包括鼻咽喉，气管)':'[MHSLOC20]','前列腺':'[MHSLOC22]'}

SHEETS_PATH = "..\sheets"

def revert(org):
    org_revert = {}
    for key in org:
        org_revert.setdefault(org[key], set())
        org_revert[org[key]].add(key)
    
    return org_revert

def data1(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        tmp_keys = deepcopy(keys)
        if ws[tmp_keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[tmp_keys['[Subject]']+str(row)].value == None:
            continue
                
        Subject = ws[tmp_keys['[Subject]']+str(row)].value
        
        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(row, {})

        tmp_keys.pop(r'{change}')
        tmp_keys.pop('[Subject]')

        for key in tmp_keys:
            data_ws[Subject][row].update({key:ws[tmp_keys[key]+str(row)].value})

    return data_ws

def data2(ws, keys):
    data_ws = {}
    if "非靶病灶" in ws.title:
        LOC = "[NTLLOC]"
    else:
        LOC = "[TLLOC]"
    for row in range(2, ws.max_row+1):

        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
                
        Subject = ws[keys['[Subject]']+str(row)].value
        InstanceName = ws[keys['[InstanceName]']+str(row)].value
        if InstanceName != "肿瘤评估-筛选期":
            continue
        TLLOC = ws[keys['%s' %LOC]+str(row)].value
        
        data_ws.setdefault(Subject, {})
        data_ws[Subject].setdefault(row, {})
        data_ws[Subject][row].update({'%s' %LOC: TLLOC})

    return data_ws

def Diagnosecheck(data_ws1, data_ws2, data_ws3, ws1):
    ws1.insert_cols(1)
    ws1['A1'].value = '肿瘤在病理学诊断检查结果'
    SLOC2LLOC = revert(LLOC2SLOC)

    for id in data_ws1:
        pid1 = data_ws1[id]
        # id_error = 0
        # if id not in data_ws2 and id not in data_ws3:
        #     id_error = 1

        for row_ws1 in pid1:
            check_error = 0
            msg = ''
            rpid1 = pid1[row_ws1]
            LLOC_set = set()

            if id not in data_ws2 and id not in data_ws3:
                # id_error = 1 
                msg = "Error: 该患者 %s 未见肿瘤评估结果" %id
                mark(ws1, 'A', row_ws1, msg)
                continue

            if id in data_ws2:
                pid2 = data_ws2[id]
                for row_ws2 in pid2:
                    rpid2 = pid2[row_ws2]
                    LLOC_set.add(rpid2['[TLLOC]'])

            if id in data_ws3:                
                pid3 = data_ws3[id]
                for row_ws3 in pid3:
                    rpid3 = pid3[row_ws3]
                    LLOC_set.add(rpid3['[NTLLOC]'])

            for MHSLOC in rpid1:
                rsg = ''
                if rpid1[MHSLOC] in ['肺', 1]:
                    if LLOC_set.intersection(SLOC2LLOC[MHSLOC]):
                        continue
                    else:
                        check_error = 1
                        rsg = "Error: 该患者 {0} {1} 核查失败".format(id, SLOC2LLOC[MHSLOC])
                
                msg = message(msg, rsg)
                    
            # if id_error:
            #     msg = "Error: 该患者 %s 未见肿瘤评估结果" %id
            if check_error:
                pass
            else:
                msg = "Info: Success"
            mark(ws1, 'A', row_ws1, msg)

def Tumorcheck(data_ws1, data_ws2, ws2):
    ws1.insert_cols(1)
    ws1['A1'].value = '肿瘤在病理学诊断检查结果'

    for id in data_ws2:
        pid2 = data_ws2[id]
        id_error = 0
        if id not in data_ws1:
            id_error = 1
        
        for row_ws2 in pid2:
            rpid2 = pid2[row_ws2]
            msg = ''
            check_error = 0
            check_warn = 0

            for value in rpid2.values():
                if value in ['食管', '腹膜', '其他', None]:
                    check_warn = 1
                else:
                    check_key = LLOC2SLOC[value]

            if not id_error and not check_warn:
                pid1 = data_ws1[id]
                for row_ws1 in pid1:
                    rpid1 = pid1[row_ws1]

                    if rpid1[check_key] not in [1, '肺']:
                        check_error = 1
                        msg = 'Error: 该患者 {0} {1} 未见病理学诊断结果'.format(id, value)

            if id_error:
                msg = "Error: 该患者 %s 未见病理学诊断结果"
            elif check_warn:
                msg = "Warn: 该患者 %s 检查内容为 %s" %(id, value)
            elif check_error:
                pass
            else:
                msg = "Info: Success"
            mark(ws2, 'A', row_ws2, msg)

def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(filename):
    for file in get_files():
        if filename in file:
            return file

if __name__ == "__main__":
    Tumor = get_a_file("KN046-301_病理学诊断和病灶")
    Tumor_path = os.path.join(SHEETS_PATH, Tumor)
    Tumor_pathlist = Tumor_path.split('.xlsx')
    wbsavepath = ''.join([''.join([Tumor_pathlist[0], '_checkout']), '.xlsx'])

    sheet1 = 'MHDIAG|鳞状非小细胞肺癌病理学诊断'
    sheet2 = 'TUTL|肿瘤评价-靶病灶（RECIST 1.1）'
    sheet3 = 'TUNTL|肿瘤评价-非靶病灶（RECIST 1.1）'
    
    SLOC2LLOC = revert(LLOC2SLOC)
    try:
        wb = openpyxl.load_workbook(Tumor_path)
        ws1 = wb[sheet1]
        ws2 = wb[sheet2]
        ws3 = wb[sheet3]

        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)
        keys3 = findkeyscolumn(ws3, keys3list)

        data_ws1 = data1(ws1, keys1)
        data_ws2 = data2(ws2, keys2)
        data_ws3 = data2(ws3, keys3)
        
        Tumorcheck(data_ws1, data_ws2, ws2)
        Tumorcheck(data_ws1, data_ws3, ws3)
        Diagnosecheck(data_ws1, data_ws2, data_ws3, ws1)
        
        wb.save(wbsavepath)

    finally:
        wb.close()