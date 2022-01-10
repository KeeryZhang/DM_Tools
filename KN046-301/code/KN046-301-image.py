#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy
import re

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message, parse_dmy

keyslist = [r'{change}','[Subject]', '[InstanceName]']

SHEETS_PATH = "..\sheets"


def data(ws, keys):
    data_ws = {}

    for row in range(2, ws.max_row+1):
        if (r'{change}' in keys and ws[keys[r'{change}']+str(row)].value == 'deleted') or (ws[keys['[Subject]']+str(row)].value is None):
            continue

        subject = ws[keys['[Subject]']+str(row)].value
        InstanceName = ws[keys['[InstanceName]']+str(row)].value
        data_ws.setdefault(subject, {'CC':{}, 'NCC':{}})

        if "筛选期" in InstanceName:
            result = ['0']
        else:
            result = re.findall(r'\d+', InstanceName)

        if "EOT" in InstanceName or "安全性随访" in InstanceName or result == []:
            continue
        elif "-CC" in InstanceName:
            data_ws[subject]['CC'].setdefault(int(result[0]), row)
        else:
            data_ws[subject]['NCC'].setdefault(int(result[0]), row)

    return data_ws

def imagecheck(data_ws, ws):
    ws.insert_cols(1)
    ws['A1'].value = '访视名称检查'

    for subject in data_ws:
        pid = data_ws[subject]
        for cc in pid:
            cpid = pid[cc]
            if cpid == {}:
                continue
            cpid_sort = sorted(cpid.items(), key=lambda time:time[0])
            n = 0
            for i in cpid_sort:
                expe = 6 * n
                if i[0] == expe:
                    n += 1
                    continue
                elif i[0] < expe:
                    continue
                else:
                    j, k = divmod(i[0], 6)
                    if j == n:
                        l = j + 1
                    else:
                        l = j
                    miss = ' '.join(str(6*x) for x in range(n, l))
                    msg = "Error: 第{}周访视缺失".format(miss)
                    mark(ws, 'A', i[1], msg)
                    if k != 0:
                        n = j + 1
                    else:
                        n = j                    

def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file

if __name__ == "__main__":
    files = get_files()
    image = get_a_file(files, "影像学检查")
    image_path = os.path.join(SHEETS_PATH, image)
    image_sheet = r'IME|影像学检查'
    image_pathlist = image_path.split('.xlsx')
    wbsavepath = ''.join([''.join([image_pathlist[0], '_checkout']), '.xlsx'])

    try:
        wb = openpyxl.load_workbook(image_path)
        ws = wb[image_sheet]
                       
        keys = findkeyscolumn(ws, keyslist)
        data_ws = data(ws, keys)
        imagecheck(data_ws, ws)

        wb.save(wbsavepath)
    finally:
        wb.close()