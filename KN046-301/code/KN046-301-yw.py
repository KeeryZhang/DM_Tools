#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.descriptors import base
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
from copy import deepcopy

sys.path.append("..\..")

from tool_lib.utils import mark, findkeyscolumn, exist, message

keys1list = ['筛选号', '访视名称', '是否跳过访视', '当前使用药物']
keys2list = [r'{Change}','[Subject]', '[InstanceName]', '[EXIDEN]', '[EXYN]']

SHEETS_PATH = "..\sheets"

def data1(ws1, keys1):
    data_ws = {}
    for row in range(7, ws1.max_row+1):
        subject = ws1[keys1['筛选号']+str(row)].value
        instance_raw = ws1[keys1['访视名称']+str(row)].value
        drug = ws1[keys1['当前使用药物']+str(row)].value
        skip = ws1[keys1['是否跳过访视']+str(row)].value

        instancelist = instance_raw.split('D')
        instance = instancelist[0]

        drugset = set()
        if drug != None:
            druglist = drug.split(';')
            for dr in druglist:
                if dr == '':
                    continue
                name, code = dr.split(':')
                if name == '无编号药物':
                    continue
                drugset.add(code.split('S')[1])
        data_ws.setdefault(subject, {})
        data_ws[subject].setdefault(instance, {'row': row, 'drug': drugset, 'skip': skip})
    return data_ws


def data2(ws2, keys2):
    data_ws = {}
    for row in range(2, ws2.max_row+1):
        if r'{change}' in keys2 and ws2[keys2[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws2[keys2['[Subject]']+str(row)].value == None:
            continue

        subject = ws2[keys2['[Subject]']+str(row)].value
        instance = ws2[keys2['[InstanceName]']+str(row)].value
        EXIDEN = ws2[keys2['[EXIDEN]']+str(row)].value
        EXYN = ws2[keys2['[EXYN]']+str(row)].value

        EXIDENset = set()
        if EXIDEN != None:
            EXIDENlist = EXIDEN.split(';')
            for line in EXIDENlist:
                EXIDENset.add(line)
        
        data_ws.setdefault(subject, {})
        data_ws[subject].setdefault(instance, {'row': row, 'drug': EXIDENset, 'EXYN': EXYN})
    return data_ws


def visitcheck(data_ws1, data_ws2, ws1):
    ws1.insert_cols(1)
    ws1['A6'].value = 'visit->研究药物给药核查结果'

    for id in data_ws1:
        pid_ws1 = data_ws1[id]
        if id not in data_ws2:
            iderror = True
        else:
            pid_ws2 = data_ws2[id]
            iderror = False

        for instance in pid_ws1:
            ipid_ws1 = pid_ws1[instance]
            msg = ''
            if 'C' not in instance:
                visitpass = True
            else:
                visitpass = False

            instanceerror = False

            if not visitpass and not iderror:    
                if instance not in pid_ws2:                    
                    instanceerror = True
                else:
                    ipid_ws2 = pid_ws2[instance]
                    instanceerror = False

            if not visitpass and not iderror and not instanceerror:
                if ipid_ws2['EXYN'] == "否" and ipid_ws1['drug'] == set():
                    exynpass = True
                else:
                    exynpass = False

            if visitpass:
                rsg = 'Info: 该行访视为 {} 无需匹配'.format(instance)
            elif iderror:
                rsg = 'Error: 该受试者{}信息在研究药物给药页面不存在'.format(id)
            elif instanceerror and ipid_ws1['skip'] == 'Y' and ipid_ws1['drug'] == set():
                rsg = 'Info: 该行跳过访视，访视 {} 在研究药物给药页面不存在'.format(instance)
            elif not instanceerror and ipid_ws1['skip'] == 'Y' and ipid_ws1['drug'] == set() and ipid_ws2['drug'] == set():
                rsg = 'Info: 该行跳过访视，药物记录在两边均为空'
            elif instanceerror:
                rsg = 'Error: 该受试者{}的访视{}信息在研究药物给药页面不存在'.format(id, instance)
            elif exynpass:
                rsg = 'Info: 该行未给药且使用药物为空'
            else:
                if ipid_ws1['drug'] == ipid_ws2['drug']:
                    rsg = 'Info: 该行使用药物匹配成功'
                else:
                    ws1diffws2 = ipid_ws1['drug'].difference(ipid_ws2['drug'])
                    ws2diffws1 = ipid_ws2['drug'].difference(ipid_ws1['drug'])
                    rsg = 'Error: 该行使用药物匹配失败，受试者 {0} 访视 {1}'.format(id, instance)            
                    if len(ws1diffws2) != 0:
                        rsg += ' visit页面多出 {}'.format(' '.join(str(x) for x in ws1diffws2))
                    if len(ws2diffws1) != 0:
                        rsg += ' 给药页面多出 {}'.format(' '.join(str(x) for x in ws2diffws1))

            msg = message(msg, rsg)
            mark(ws1, 'A', ipid_ws1['row'], msg)


def gycheck(data_ws1, data_ws2, ws1):
    ws1.insert_cols(1)
    ws1['A6'].value = '研究药物给药->visit核查结果'

    for id in data_ws1:
        pid_ws1 = data_ws1[id]
        if id not in data_ws2:
            iderror = True
        else:
            pid_ws2 = data_ws2[id]
            iderror = False

        for instance in pid_ws1:
            ipid_ws1 = pid_ws1[instance]
            msg = ''
            if 'C' not in instance:
                visitpass = True
            elif 'CC' in instance:
                visitpass = True
            else:
                visitpass = False

            if not visitpass and not iderror:
                if instance not in pid_ws2:
                    instanceerror = True
                else:
                    ipid_ws2 = pid_ws2[instance]
                    instanceerror = False

            if not visitpass and not iderror and not instanceerror:
                if ipid_ws1['EXYN'] == "否" and ipid_ws2['drug'] == set():
                    exynpass = True
                else:
                    exynpass = False

            if visitpass:
                rsg = 'Info: 该行访视为 {} 无需匹配'.format(instance)
            elif iderror:
                rsg = 'Error: 该受试者{}信息在受试者页面不存在'.format(id)
            elif ipid_ws1['EXYN'] == "否" and instanceerror:
                rsg = 'Info: 该行未给药且受试者页面不存在访视 {}'.format(instance)
            elif instanceerror:
                rsg = 'Error: 该受试者{}的访视{}信息在受试者页面不存在'.format(id, instance)
            elif exynpass:
                rsg = 'Info: 该行未给药且使用药物为空'            
            else:
                if ipid_ws1['drug'] == ipid_ws2['drug']:
                    rsg = 'Info: 该行使用药物匹配成功'
                else:
                    rsg = 'Error: 该行药物匹配失败，已在IRT受试者页面记录'
                    # ws1diffws2 = ipid_ws1['drug'].difference(ipid_ws2['drug'])
                    # ws2diffws1 = ipid_ws2['drug'].difference(ipid_ws1['drug'])
                    # rsg = 'Error: 该行使用药物匹配失败，受试者 {0} 访视 {1}'.format(id, instance)            
                    # if len(ws1diffws2) != 0:
                    #     rsg += ' 给药页面多出 {}'.format(' '.join(str(x) for x in ws1diffws2))
                    # if len(ws2diffws1) != 0:
                    #     rsg += ' visit页面多出 {}'.format(' '.join(str(x) for x in ws2diffws1))

            msg = message(msg, rsg)
            mark(ws1, 'A', ipid_ws1['row'], msg)


def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    # parser.add_argument("--visit", default=r'Copy of Subject_visit_report_KN046-3011629948672870.xlsx', help="Please add AE file name")
    # parser.add_argument("--gy", default=r'KN046-301_研究药物给药.xlsx', help="Please set sheet name of ae")
    parser.add_argument("--ssz", default=r'受试者', help="Please set sheet name of cb")
    parser.add_argument("--ywgy", default=r'EX|研究药物给药', help="Please set sheet name of cb")

    files = get_files()
    visit = get_a_file(files, "Subject_visit_report_KN046-301")
    gy = get_a_file(files, "KN046-301_研究药物给药")
    args = parser.parse_args()

    visit_path = os.path.join(SHEETS_PATH, visit)
    gy_path = os.path.join(SHEETS_PATH, gy)

    ssz_sheet = args.ssz
    ywgy_sheet = args.ywgy

    visit_pathlist = visit_path.split('.xlsx')
    gy_pathlist = gy_path.split('.xlsx')

    wb1savepath = ''.join([''.join([visit_pathlist[0], '_checkout']), '.xlsx'])
    wb2savepath = ''.join([''.join([gy_pathlist[0], '_checkout']), '.xlsx'])
    try:
        wb1 = openpyxl.load_workbook(visit_path)
        ws1 = wb1[ssz_sheet]
        
        wb2 = openpyxl.load_workbook(gy_path)
        ws2 = wb2[ywgy_sheet]
               
        keys1 = {'筛选号': 'C', '访视名称': 'L', '当前使用药物': 'R', '是否跳过访视': 'O'}
        keys2 = findkeyscolumn(ws2, keys2list)

        data_ws1 = data1(ws1, keys1)
        data_ws2 = data2(ws2, keys2)
        
        visitcheck(data_ws1, data_ws2, ws1)
        gycheck(data_ws2, data_ws1, ws2)

        wb1.save(wb1savepath)
        wb2.save(wb2savepath)

    finally:
        wb1.close()
        wb2.close()