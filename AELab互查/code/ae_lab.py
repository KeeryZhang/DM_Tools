#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from datetime import datetime
import os
import sys
import argparse
import copy

sys.path.append("..\..")
from tool_lib.outs.Codelist_map import Codelist_map as Codelist_map
from tool_lib.outs.Coderevert_map import Coderevert_map as Coderevert_map

M2m = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
keys1list = [r'{change}','[Subject]','[AEYN]','[AETERM_PT]','[AESTDAT_RAW]','[AEENDAT_RAW]','[AEGRDAT_RAW]','[AEOUT]']
keys2list = [r'{change}','[Subject]','[RecordDate]','[AnalyteName]','[AnalyteValue]','[LabFlag]','[ClinSigValue]']

# Codelist_map = {('RBC', '-'): ['Red blood cell count decreased', 'Anaemia'], ('HGB', '-'): ['Anaemia'], ('HCT', '-'): ['Haematocrit decreased'], ('WBC', '+'): ['White blood cell count increased'], ('WBC', '-'): ['White blood cell count decreased'], ('NEUT', '+'): ['Neutrophil count increased'], ('NEUT', '-'): ['Neutrophil count decreased'], ('EOS', '+'): ['Eosinophil count increased'], ('EOS', '-'): ['Eosinophil count decreased'], ('BASO', '+'): ['Basophil count increased'], ('BASO', '-'): ['Basophilopenia'], ('LYM', '+'): ['Lymphocyte count increased'], ('LYM', '-'): ['Lymphocyte count decreased'], ('MONO', '+'): ['Monocyte count increased', 'Monocyte percentage increased'], ('MONO', '-'): ['Lymphocyte count decreased'], ('PLAT', '+'): ['Platelet count increased'], ('PLAT', '-'): ['Platelet count decreased', 'Thrombocytopenia'], ('BILI', '+'): ['Blood bilirubin increased'], ('ALT', '+'): ['Alanine aminotransferase increased'], ('AST', '+'): ['Aspartate aminotransferase increased'], ('GGT', '+'): ['Gamma-glutamyltransferase increased'], ('ALP', '+'): ['Blood alkaline phosphatase increased'], ('ALB', '-'): ['Hypoalbuminaemia', 'Blood albumin decreased'], ('PROT', '-'): ['Protein total decreased'], ('LDH', '+'): ['Blood lactate dehydrogenase increased'], ('UREA', '+'): ['Blood urea increased'], ('CREAT', '+'): ['Blood creatinine increased'], ('SODIUM', '+'): ['Hypernatraemia'], ('SODIUM', '-'): ['Hyponatraemia', 'Blood sodium decreased'], ('K', '+'): ['Hyperkalaemia', 'Blood potassium increased'], ('K', '-'): ['Hypokalaemia', 'Blood potassium decreased'], ('CL', '+'): ['Hyperchloraemia'], ('CL', '-'): ['Hypochloraemia', 'Blood chloride decreased'], ('MG', '+'): ['Hypermagnesaemia'], ('MG', '-'): ['Hypomagnesaemia', 'Blood magnesium decreased'], ('CA', '+'): ['Blood calcium increased', 'Hypercalcaemia'], ('CA', '-'): ['Blood calcium decreased', 'Calcium ionised decreased', 'Hypocalcaemia'], ('PHOS', '+'): ['Blood phosphorus increased', 'Hyperphosphataemia'], ('PHOS', '-'): ['Hypophosphataemia', 'Blood phosphorus decreased'], ('AMYLASE', '+'): ['Amylase increased'], ('GLUC_FAST', '+'): ['Hyperglycaemia', 'Blood glucose increased'], ('GLUC_FAST', '-'): ['Hypoglycaemia'], ('BILDIR', '+'): ['Bilirubin conjugated increased'], ('CK', '+'): ['Blood creatine phosphokinase increased'], ('UREAN', '+'): ['Blood urea increased'], ('UREAN', '-'): ['Blood urea decreased'], ('PT', '+'): ['Prothrombin time prolonged'], ('INR', '+'): ['International normalised ratio increased'], ('T3', '-'): ['Tri-iodothyronine decreased'], ('T3FR', '+'): ['Tri-iodothyronine free increased'], ('T3FR', '-'): ['Tri-iodothyronine free decreased'], ('T4FR', '+'): ['Thyroxine free increased'], ('T4FR', '-'): ['Thyroxine free decreased'], ('TSH', '+'): ['Blood thyroid stimulating hormone increased'], ('TSH', '-'): ['Blood thyroid stimulating hormone decreased'], ('CKMB', '+'): ['Creatine kinase MB increased'], ('CRP', '+'): ['C-reactive protein increased']}


# Coderevert_map = {'Red blood cell count decreased': [('RBC', '-')], 'Anaemia': [('RBC', '-'), ('HGB', '-')], 'Haematocrit decreased': [('HCT', '-')], 'White blood cell count increased': [('WBC', '+')], 'White blood cell count decreased': [('WBC', '-')], 'Neutrophil count increased': [('NEUT', '+')], 'Neutrophil count decreased': [('NEUT', '-')], 'Eosinophil count increased': [('EOS', '+')], 'Eosinophil count decreased': [('EOS', '-')], 'Basophil count increased': [('BASO', '+')], 'Basophilopenia': [('BASO', '-')], 'Lymphocyte count increased': [('LYM', '+')], 'Lymphocyte count decreased': [('LYM', '-'), ('MONO', '-')], 'Monocyte count increased': [('MONO', '+')], 'Monocyte percentage increased': [('MONO', '+')], 'Platelet count increased': [('PLAT', '+')], 'Platelet count decreased': [('PLAT', '-')], 'Thrombocytopenia': [('PLAT', '-')], 'Blood bilirubin increased': [('BILI', '+')], 'Alanine aminotransferase increased': [('ALT', '+')], 'Aspartate aminotransferase increased': [('AST', '+')], 'Gamma-glutamyltransferase increased': [('GGT', '+')], 'Blood alkaline phosphatase increased': [('ALP', '+')], 'Hypoalbuminaemia': [('ALB', '-')], 'Blood albumin decreased': [('ALB', '-')], 'Protein total decreased': [('PROT', '-')], 'Blood lactate dehydrogenase increased': [('LDH', '+')], 'Blood urea increased': [('UREA', '+'), ('UREAN', '+')], 'Blood creatinine increased': [('CREAT', '+')], 'Hypernatraemia': [('SODIUM', '+')], 'Hyponatraemia': [('SODIUM', '-')], 'Blood sodium decreased': [('SODIUM', '-')], 'Hyperkalaemia': [('K', '+')], 'Blood potassium increased': [('K', '+')], 'Hypokalaemia': [('K', '-')], 'Blood potassium decreased': [('K', '-')], 'Hyperchloraemia': [('CL', '+')], 'Hypochloraemia': [('CL', '-')], 'Blood chloride decreased': [('CL', '-')], 'Hypermagnesaemia': [('MG', '+')], 'Hypomagnesaemia': [('MG', '-')], 'Blood magnesium decreased': [('MG', '-')], 'Blood calcium increased': [('CA', '+')], 'Hypercalcaemia': [('CA', '+')], 'Blood calcium decreased': [('CA', '-')], 'Calcium ionised decreased': [('CA', '-')], 'Hypocalcaemia': [('CA', '-')], 'Blood phosphorus increased': [('PHOS', '+')], 'Hyperphosphataemia': [('PHOS', '+')], 'Hypophosphataemia': [('PHOS', '-')], 'Blood phosphorus decreased': [('PHOS', '-')], 'Amylase increased': [('AMYLASE', '+')], 'Hyperglycaemia': [('GLUC_FAST', '+')], 'Blood glucose increased': [('GLUC_FAST', '+')], 'Hypoglycaemia': [('GLUC_FAST', '-')], 'Bilirubin conjugated increased': [('BILDIR', '+')], 'Blood creatine phosphokinase increased': [('CK', '+')], 'Blood urea decreased': [('UREAN', '-')], 'Prothrombin time prolonged': [('PT', '+')], 'International normalised ratio increased': [('INR', '+')], 'Tri-iodothyronine decreased': [('T3', '-')], 'Tri-iodothyronine free increased': [('T3FR', '+')], 'Tri-iodothyronine free decreased': [('T3FR', '-')], 'Thyroxine free increased': [('T4FR', '+')], 'Thyroxine free decreased': [('T4FR', '-')], 'Blood thyroid stimulating hormone increased': [('TSH', '+')], 'Blood thyroid stimulating hormone decreased': [('TSH', '-')], 'Creatine kinase MB increased': [('CKMB', '+')], 'C-reactive protein increased': [('CRP', '+')]}


def findkeyscolumn(ws, keyslist):
    keys = {}
    for column in range(1, ws.max_column+1):
        row_letter = get_column_letter(column)
        for key in keyslist:
            if key in ws[row_letter+'1'].value:
                keys.setdefault(key, row_letter)
                keyslist.remove(key)
                break
    return keys


def parse_dmy(s):
    if s == None:
        s = 'UN UNK 0000'
    day_s,mon_s,year_s=s.split(' ')
    if day_s == 'UN':
        day_s = '1'
    if mon_s == 'UNK':
        mon_s = 'JAN'
    if year_s == '0000':
        year_s = '1970'
    return datetime(int(year_s),int(M2m[mon_s]),int(day_s))


def mark(ws, col, row, msg):
    ws[col+str(row)] = msg
    return


def data1(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        if ws[keys[r'{change}']+str(row)] == 'deleted' or ws[keys['[AEYN]']+str(row)] == '否' or ws[keys['[AEYN]']+str(row)] == None:
            continue
        patientId = ws[keys['[Subject]']+str(row)].value
        AE_PT = ws[keys['[AETERM_PT]']+str(row)].value
        AEOUT = ws[keys['[AEOUT]']+str(row)].value

        AEST = ws[keys['[AESTDAT_RAW]']+str(row)].value
        AEEN = ws[keys['[AEENDAT_RAW]']+str(row)].value
        AEGR = ws[keys['[AEGRDAT_RAW]']+str(row)].value
        ST = parse_dmy(AEST)
        EN = parse_dmy(AEEN)
        GR = parse_dmy(AEGR)

        data_ws.setdefault(patientId, {})
        data_ws[patientId].setdefault(AE_PT, {})
        data_ws[patientId][AE_PT].setdefault(ST, {})
        data_ws[patientId][AE_PT][ST].setdefault(row, {'EN':EN, 'GR':GR, 'AEOUT':AEOUT})
    return data_ws


def data2(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        if ws[keys[r'{change}']+str(row)].value == 'deleted':
            continue
        if ws[keys['[Subject]']+str(row)].value == None:
            continue
        patientId = ws[keys['[Subject]']+str(row)].value
        RecordDate = ws[keys['[RecordDate]']+str(row)].value
        AnalyteName = ws[keys['[AnalyteName]']+str(row)].value
        AnalyteValue = ws[keys['[AnalyteValue]']+str(row)].value
        LabFlag = ws[keys['[LabFlag]']+str(row)].value
        CS = ws[keys['[ClinSigValue]']+str(row)].value

        data_ws.setdefault(patientId, {})
        data_ws[patientId].setdefault(AnalyteName, {})
        data_ws[patientId][AnalyteName].setdefault(row, {'RecordDate':RecordDate, 'AnalyteValue':AnalyteValue, 'LabFlag':LabFlag, 'CS':CS})
    return data_ws


def lab2ae(data_ws1, data_ws2, ws2):
    ws2.insert_cols(1)
    ws2['A1'].value = 'YD Comment'

    for id in data_ws2:
        pid = data_ws2[id]
        for AnalyteName in pid:            
            STapid_ws1 = {}
            apid = pid[AnalyteName]
            sortedapid = sorted(apid.items(), key = lambda time:time[1]['RecordDate'])
            AE_event = False
            row_ws1 = None
            AETERM_PT = None
            for row_ws2 in sortedapid:
                rs = row_ws2[1]
                if rs['LabFlag'] in ['+', '-'] and rs['CS'] == 'Clinically Significant':
                    if id not in data_ws1.keys():
                        msg = 'Error:该患者在AE页面无记录'
                        mark(ws2, 'A', row_ws2[0], msg)
                        continue       
                    
                    if (AnalyteName, rs['LabFlag']) not in Codelist_map:
                        msg = 'Warn:该行分析物名称无对应不良事件名称，请核查'
                        mark(ws2, 'A', row_ws2[0], msg)
                        continue 

                    if not AE_event:
                        codelist = Codelist_map[(AnalyteName, rs['LabFlag'])]
                        codecheck = False
                        for code in codelist:
                            if code in data_ws1[id]:
                                codecheck = True
                                AETERM_PT = code
                                break
                        
                        if not codecheck:
                            msg = 'Error:该患者在AE页面无{}记录'.format(AnalyteName)
                            mark(ws2, 'A', row_ws2[0], msg)
                            continue                         
                    
                        apid_ws1 = data_ws1[id][AETERM_PT]
                        if rs['RecordDate'] in apid_ws1:
                            STapid_ws1_pre = copy.deepcopy(apid_ws1[rs['RecordDate']])
                            AE_event = True
                            msg = 'Info:该行为AE开始，在AE页面有相应记录'
                            mark(ws2, 'A', row_ws2[0], msg)
                            STapid_ws1 = sorted(STapid_ws1_pre.items(), key = lambda time:time[1]['GR'])
                            raw_ws1 = STapid_ws1[0]
                            continue

                        else:
                            msg = 'Error:该行有AE发生，但未在AE页面找到相应记录'
                            mark(ws2, 'A', row_ws2[0], msg)
                            continue
                                        
                    else:
                        checkline = False
                        for line in STapid_ws1:
                            if rs['RecordDate'] == line[1]['GR']:
                                raw_ws1 = line
                                msg = 'Info:该行对应AE页面第{}行，发生级别变化'.format(raw_ws1[0])
                                mark(ws2, 'A', row_ws2[0], msg)
                                checkline = True
                                break
                        if not checkline:
                            msg = 'Info:该行处于AE发生中，无AE页面对应，未检测到级别变化'
                            mark(ws2, 'A', row_ws2[0], msg)
                            continue

                if rs['LabFlag'] == None and rs['AnalyteValue'] == None:
                    msg = 'Info:该行未录入数据'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == None and rs['AnalyteValue'] in ['#NA', '#ND']:
                    msg = 'Info:该项目未检测'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == None:
                    msg = 'Error:实验室检测范围缺失'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == '0' and not AE_event:
                    msg = 'Info:该行数值在正常范围内，无需核查'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] == '0' and AE_event:
                    EN = raw_ws1[1]['EN']
                    if rs['RecordDate'] == EN:
                        msg = 'Info:该检测值回归正常'
                    else:
                        msg = 'Error:该行应为AE结束日期，但在AE页面第{}行无对应'.format(raw_ws1[0])
                    raw_ws1 = None
                    mark(ws2, 'A', row_ws2[0], msg)
                    AE_event = False
                    continue
                elif rs['LabFlag'] in ['+', '-'] and rs['CS'] == 'Not Clinically Significant' and not AE_event:
                    msg = 'Info:该行NCS，无需核查'
                    mark(ws2, 'A', row_ws2[0], msg)
                    continue
                elif rs['LabFlag'] in ['+', '-'] and rs['CS'] == 'Not Clinically Significant' and AE_event:
                    EN = raw_ws1[1]['EN']
                    if rs['RecordDate'] == EN:
                        msg = 'Info:该行NCS回归正常'
                    else:
                        msg = 'Error:该行应为AE结束日期，但在AE页面第{}行无对应'.format(raw_ws1[0])
                    raw_ws1 = None
                    mark(ws2, 'A', row_ws2[0], msg)
                    AE_event = False
                    continue

    return ws2


def ae2lab(data_ws1, data_ws2, ws1):
    ws1.insert_cols(1)
    ws1['A1'].value = 'YD Comment'
    
    for id in data_ws1:
        pid = data_ws1[id]
        for AE_PT in pid:
            apid = pid[AE_PT]
            if AE_PT in Coderevert_map:
                AnalyteNames = Coderevert_map[AE_PT]
            else:
                AnalyteNames = None
            for ST in apid:                
                sapid = apid[ST]
                sortedsapid = sorted(sapid.items(), key = lambda time:time[1]['GR'])
                for row_ws1 in sortedsapid:
                    msg = ''
                    if id not in data_ws2.keys():
                        msg = 'Error:该患者在Lab页面无记录'
                        mark(ws1, 'A', row_ws1[0], msg)
                        continue
                    pid_ws2 = data_ws2[id]
                    
                    AnalyteNamecheck = False
                    if AnalyteNames == None:
                        msg = 'Warn:该行不良事件{}无对应分析物，需手动核查'.format(AE_PT)
                        mark(ws1, 'A', row_ws1[0], msg)
                        continue
                    for An_Flag in AnalyteNames:
                        if An_Flag[0] in pid_ws2:
                            AnalyteNamecheck = True
                            apid_ws2 = pid_ws2[An_Flag[0]]
                            STcheck = False
                            GRcheck = False
                            ENcheck = False
                            GR = False
                            EN = False
                            if row_ws1[1]['GR'] != datetime(1970, 1, 1):
                                GR = True
                            if row_ws1[1]['EN'] != datetime(1970, 1, 1):
                                EN = True
                            for row_ws2 in apid_ws2:
                                rapid_ws2 = apid_ws2[row_ws2]
                                if rapid_ws2['RecordDate'] == ST:
                                    STcheck = True
                                    msg = 'Info:该AE记录开始日期与Lab页面第{}行匹配成功'.format(row_ws2)
                                elif GR and rapid_ws2['RecordDate'] == row_ws1[1]['GR']:
                                    GRcheck = True
                                    msg = '\n'.join([msg, 'Info:该AE记录级别变化日期与Lab页面第{}行匹配成功'.format(row_ws2)])
                                elif EN and rapid_ws2['RecordDate'] == row_ws1[1]['EN']:
                                    ENcheck = True
                                    msg = '\n'.join([msg, 'Info:该AE记录结束日期与Lab页面第{}行匹配成功'.format(row_ws2)])
                            
                            if not STcheck:
                                msg = '\n'.join([msg, 'Error:该AE记录开始日期匹配失败'])
                            if GR and not GRcheck:
                                msg = '\n'.join([msg, 'Error:该AE记录级别变化日期匹配失败'])
                            if EN and not ENcheck:
                                msg = '\n'.join([msg, 'Error:该AE记录结束日期匹配失败'])
                            mark(ws1, 'A', row_ws1[0], msg)

                    if not AnalyteNamecheck:        
                        msg = 'Error:该患者在Lab页面无{}对应检查项'.format(AE_PT)
                        mark(ws1, 'A', row_ws1[0], msg)
                        continue
                         

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ae", default=r'E:\python\DM_Tool\20210726\CIBI308C303_AE_20210726.xlsx', help="Please add AE file full path")
    parser.add_argument("--aesheet", default=r'AE001_1|不良事件-不包括输液反应及免疫相关不良事件', help="Please set sheet name of ae")
    parser.add_argument("--lab", default=r'E:\python\DM_Tool\20210726\CIBI308C303_Lab_20210726.xlsx', help="Please add CB file full path")
    parser.add_argument("--labsheet", default=r'LAB|实验室检测', help="Please set sheet name of cb")
    parser.add_argument("--flow", default="all", help="Please state the flow you need to run")

# Test parts
    # parser.add_argument("--ae", default=r'C:\Users\zhangk22\Documents\python\project3\AEtest.xlsx', help="Please add AE file full path")
    # parser.add_argument("--aesheet", default=r'Sheet1', help="Please set sheet name of ae")
    # parser.add_argument("--lab", default=r'C:\Users\zhangk22\Documents\python\project3\labtest.xlsx', help="Please add CB file full path")
    # parser.add_argument("--labsheet", default=r'Sheet1', help="Please set sheet name of cb")
    # parser.add_argument("--flow", default="all", help="Please state the flow you need to run")

    args = parser.parse_args()

    ae_path = args.ae
    ae_sheet = args.aesheet
    lab_path = args.lab
    lab_sheet = args.labsheet
    flow = args.flow

    ae_pathlist = ae_path.split('.')
    lab_pathlist = lab_path.split('.')

    wb1savepath = '.'.join([''.join([ae_pathlist[0], '_checkout']), ae_pathlist[1]])
    wb2savepath = '.'.join([''.join([lab_pathlist[0], '_checkout']), lab_pathlist[1]])
    try:
        wb1 = openpyxl.load_workbook(ae_path)
        ws1 = wb1.get_sheet_by_name(ae_sheet)

        wb2 = openpyxl.load_workbook(lab_path)
        ws2 = wb2.get_sheet_by_name(lab_sheet)
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2 = findkeyscolumn(ws2, keys2list)

        data_ws1 = data1(ws1, keys1)
        data_ws2 = data2(ws2, keys2)

        ws2 = lab2ae(data_ws1, data_ws2, ws2)
        ws1 = ae2lab(data_ws1, data_ws2, ws1)

        wb2.save(wb2savepath)
        wb1.save(wb1savepath)

    finally:
        wb1.close()
        wb2.close()
