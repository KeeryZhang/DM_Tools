#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

if __name__ == "__main__":
    rave_manual_Codelist = r'E:\python\project3\rave_manual_Codelist.xlsx'

    wb = openpyxl.load_workbook(rave_manual_Codelist)
    ws = wb.get_sheet_by_name(r'spec_005')

    Codelist_map = {}
    for row in range(2,ws.max_row+1):
        check = ws['D'+str(row)].value

        plus_raw = ws['E'+str(row)].value
        if plus_raw == None:
            pass
        else:
            if ',' in plus_raw:
                plus = plus_raw.split(',')
            else:
                plus = [plus_raw]
            Codelist_map.setdefault((check, '+'), plus)

        minus_raw = ws['F'+str(row)].value
        if minus_raw == None:
            pass
        else:
            if ',' in minus_raw:
                minus = minus_raw.split(',')
            else:
                minus = [minus_raw]
            Codelist_map.setdefault((check, '-'), minus)        

    wb.close()
    print(Codelist_map)
    