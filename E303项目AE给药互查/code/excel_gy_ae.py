#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
import copy


M2m = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
keys1list = ['[Subject]','[AESTDAT_RAW]','[AETCDAT_RAW]','[AEACN]','[AEACN2]','[AEACN3]']
keys2list = [r'{change}','[Subject]','[EXYN]','[EXSTDAT_RAW]']
catchlist = {'研究药物剂量暂停', '研究药物剂量停用'}


def parse_dmy(s):
    day_s,mon_s,year_s=s.split(' ')
    if day_s == 'UN':
        day_s = '1'
    if mon_s == 'UNK':
        mon_s = 'JAN'
    if year_s == '0000':
        year_s = '1970'

    return datetime(int(year_s),int(M2m[mon_s]),int(day_s))


def mark(ws, row, col, msg):
    ws[col+str(row)] = msg
    return


def findkeyscolumn(ws, keyslist):
    keys = []
    for column in range(1, ws.max_column+1):
        row_letter = get_column_letter(column)
        for key in keyslist:
            if key in ws[row_letter+'1'].value:
                keys.append(row_letter)
    return keys


def data1(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        overwrite = False
        aeacn1 = ws[keys[3]+str(row)].value
        aeacn2 = ws[keys[4]+str(row)].value
        aeacn3 = ws[keys[5]+str(row)].value

        if (aeacn1 in catchlist) or (aeacn2 in catchlist):
            patientId = ws[keys[0]+str(row)].value
            aetcdat = ws[keys[2]+str(row)].value
            if aetcdat == None:
                starttime = ws[keys[1]+str(row)].value
            else:
                starttime = aetcdat
                overwrite = True
            if starttime == None:
                starttime = 'UN UNK 0000'
            else:
                st = parse_dmy(starttime)
            data_ws.setdefault(patientId, {})
            data_ws[patientId].setdefault(row, {'st':st, 'aeacn1':[aeacn1, ""], 'aeacn2':[aeacn2, ""], 'aeacn3':[aeacn3, ""], 'overwrite':overwrite})
    return data_ws


def data2(ws, keys):
    data_ws = {}

    for row in range(2, ws.max_row+1):
        change = ws[keys[0]+str(row)].value
        if change == 'deleted':
            continue

        exyn = ws[keys[2]+str(row)].value

        if exyn == '是':
            patientId = ws[keys[1]+str(row)].value    
            er = ws[keys[3]+str(row)].value
            if er == None:
                er = 'UN UNK 0000'
            st = parse_dmy(er)
            data_ws.setdefault(patientId, [])
            data_ws[patientId].append(st)
    return data_ws


def data2_new(ws, keys):
    data_ws = {}
    for row in range(2, ws.max_row+1):
        change = ws[keys[0]+str(row)].value
        if change == 'deleted':
            continue

        exyn = ws[keys[2]+str(row)].value

        if exyn == '是':
            patientId = ws[keys[1]+str(row)].value   
            er = ws[keys[3]+str(row)].value
            if er == None:
                er = 'UN UNK 0000'
            st = parse_dmy(er)
            data_ws.setdefault(patientId,{})
            data_ws[patientId].setdefault(row, {'st':st, 'msg':'', 'diff':0})
    return data_ws


def crosscheck(data_ws1, data_ws2list, ws1, ws2list):
    ws1.insert_cols(1)
    ws1.insert_cols(1)
    ws1.insert_cols(1)

    ws1['A1'].value = "IBI308安慰剂"
    ws1['B1'].value = "奥沙利铂"
    ws1['C1'].value = "卡培他滨"

    checklist = ['aeacn1', 'aeacn2', 'aeacn3']

    for id in data_ws1:
        pid = data_ws1[id]
        for row_ws1 in pid:
            pr1 = pid[row_ws1]
            for i in range(0,len(data_ws2list)):
                timelist=[]
                check = checklist[i]
                if pr1[check][0] in catchlist:
                    data_ws2=data_ws2list[i]
                    if id in data_ws2:
                        st = pr1['st']
                        timelist = copy.deepcopy(data_ws2[id])
                        if timelist.count(st) == 0:                    
                            timelist.append(st)                            
                            timelist.sort()
                            index = timelist.index(st)
                            if st == timelist[-1]:
                                pr1[check][1] = "Info:用药页面日期早于AE开始日期"
                                if pr1['overwrite']:
                                    pr1[check][1] += "(级别变化覆盖开始日期）"
                                mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])
                            elif index == 0:
                                pr1[check][1] = "Error:AE开始日期早于用药开始日期"
                                if pr1['overwrite']:
                                    pr1[check][1] += "(级别变化覆盖开始日期）"                            
                                mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])
                            else:
                                early = timelist[index-1]
                                late = timelist[index+1]
                                diff = (late - early).days
                                if diff >= 24:
                                    pr1[check][1] = "Info:用药页面间隔时间大于等于24天"
                                    if pr1['overwrite']:
                                        pr1[check][1] += "(级别变化覆盖开始日期）"                                
                                    mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])
                                else:
                                    pr1[check][1] = "Error:用药页面间隔时间小于24天"
                                    if pr1['overwrite']:
                                        pr1[check][1] += "(级别变化覆盖开始日期）"                                
                                    mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1]) 

                        else:
                            timelist.sort()   
                            timelist.reverse()    
                            if st == timelist[0]:
                                pr1[check][1] = "Info:用药页面日期早于AE开始日期"
                                if pr1['overwrite']:
                                    pr1[check][1] += "(级别变化覆盖开始日期）"
                                mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])
                            elif st == timelist[-1]:
                                pr1[check][1] = "Error:AE开始日期早于用药开始日期"
                                if pr1['overwrite']:
                                    pr1[check][1] += "(级别变化覆盖开始日期）"                            
                                mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])
                            else:
                                index = timelist.index(st)
                                late = timelist[index-1]
                                diff = (late - st).days
                                if diff >= 24:
                                    pr1[check][1] = "Info:用药页面间隔时间大于等于24天"
                                    if pr1['overwrite']:
                                        pr1[check][1] += "(级别变化覆盖开始日期）"                                
                                    mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])
                                else:
                                    pr1[check][1] = "Error:用药页面间隔时间小于24天"
                                    if pr1['overwrite']:
                                        pr1[check][1] += "(级别变化覆盖开始日期）"                                
                                    mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])

                    else:
                        pr1[check][1] = "该患者在 "+ws2list[i].title+" 中不存在"
                        mark(ws1, row_ws1, get_column_letter(i+1), pr1[check][1])
                else:
                    continue
    return ws1


def reversecheck(data_ws1, data_ws2, ws2, ae):
    ws2.insert_cols(1)
    ws2['A1'].value = '反向核查结果'
    for id in data_ws2:
        pid = data_ws2[id]
        if len(pid) <= 1:
            for row in pid:
                pid[row]['msg'] = 'Info:该患者只有一条给药记录，请从AE界面进行核查'
                mark(ws2, row, 'A', pid[row]['msg'])
        else:
            sortedpid = sorted(pid.items(), key = lambda time:time[1]['st'])
            for i in range(1, len(sortedpid)):
                processing = False
                diff = (sortedpid[i][1]['st'] - sortedpid[i-1][1]['st']).days
                row_ws2 = sortedpid[i][0]
                pid[row_ws2]['diff'] = diff
                if diff > 24:
                    if not (id in data_ws1):
                        pid[row_ws2]['msg'] = 'Error:本行超窗为' + str(diff) + '天，AE页面该患者无暂停或停用记录'
                        mark(ws2, row_ws2, 'A', pid[row_ws2]['msg'])
                        continue
                    pid_ws1 = data_ws1[id]
                    for row_ws1 in pid_ws1:
                        aeacn = pid_ws1[row_ws1][ae][0]
                        if aeacn in catchlist:
                            st_ws1 = pid_ws1[row_ws1]['st']
                            if st_ws1 <= sortedpid[i][1]['st'] and st_ws1 >= sortedpid[i-1][1]['st']:
                                pid[row_ws2]['msg'] = 'Info:本行超窗为' + str(diff) + '天，AE页面存在符合条件的记录，对应行数' + str(row_ws1)
                                mark(ws2, row_ws2, 'A', pid[row_ws2]['msg'])
                                processing = True
                                break
                    if not processing:
                        pid[row_ws2]['msg'] = 'Error:本行超窗为' + str(diff) + '天，AE页面无符合条件的记录'
                        mark(ws2, row_ws2, 'A', pid[row_ws2]['msg'])
                else:
                    pid[row_ws2]['msg'] = 'Info:本行超窗为' + str(diff) + '天'
                    mark(ws2, row_ws2, 'A', pid[row_ws2]['msg'])
                       

    return


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--ae", default=r'E:\python\project2\ae_YD20210721.xlsx', help="Please add AE file full path")
    parser.add_argument("--aesheet", default=r'AE001_1|不良事件-不包括输液反应及免疫相关不良事件', help="Please set sheet name of ae")
    parser.add_argument("--gy", default=r'E:\python\project2\给药页面.xlsx', help="Please add CB file full path")
    parser.add_argument("--gysheet1", default=r'EX001_1|研究给药_IBI308安慰剂', help="Please set sheet name of cb")
    parser.add_argument("--gysheet2", default=r'EX001_2|研究给药_奥沙利铂', help="Please set sheet name of cb")
    parser.add_argument("--flow", default="all", help="Please state the flow you need to run")

# Test parts
    # parser.add_argument("--ae", default=r'E:\python\project2\ae_YD20210721test.xlsx', help="Please add AE file full path")
    # parser.add_argument("--aesheet", default=r'38001', help="Please set sheet name of ae")
    # parser.add_argument("--gy", default=r'E:\python\project2\给药页面test.xlsx', help="Please add CB file full path")
    # parser.add_argument("--gysheet1", default=r'38001', help="Please set sheet name of cb")
    # parser.add_argument("--gysheet2", default=r'EX001_2|研究给药_奥沙利铂', help="Please set sheet name of cb")
    # parser.add_argument("--flow", default="all", help="Please state the flow you need to run")

    args = parser.parse_args()

    ae_path = args.ae
    ae_sheet = args.aesheet
    gy_path = args.gy
    gy_sheet1 = args.gysheet1
    gy_sheet2 = args.gysheet2
    flow = args.flow

    ae_pathlist = ae_path.split('.')
    gy_pathlist = gy_path.split('.')

    wb1savepath = '.'.join([''.join([ae_pathlist[0], '_checkout']), ae_pathlist[1]])
    wb2savepath = '.'.join([''.join([gy_pathlist[0], '_checkout']), gy_pathlist[1]])
    try:
        wb1 = openpyxl.load_workbook(ae_path)
        ws1 = wb1.get_sheet_by_name(ae_sheet)

        wb2 = openpyxl.load_workbook(gy_path)
        ws2_1 = wb2.get_sheet_by_name(gy_sheet1)
        ws2_2 = wb2.get_sheet_by_name(gy_sheet2)

        ws2 = [ws2_1, ws2_2]
               
        keys1 = findkeyscolumn(ws1, keys1list)
        keys2_1 = findkeyscolumn(ws2_1, keys2list)
        keys2_2 = findkeyscolumn(ws2_2, keys2list)


        data_ws1 = data1(ws1, keys1)
        data_ws2_1 = data2_new(ws2_1, keys2_1)
        data_ws2_2 = data2_new(ws2_2, keys2_2)


        data_ws2 = [data_ws2_1, data_ws2_2]

        # ws1 = crosscheck(data_ws1, data_ws2, ws1, ws2)
        ws2_1 = reversecheck(data_ws1, data_ws2_1, ws2_1, 'aeacn1')
        ws2_2 = reversecheck(data_ws1, data_ws2_2, ws2_2, 'aeacn2')

        # wb1.save(wb1savepath)
        wb2.save(wb2savepath)

    finally:
        wb1.close()
        wb2.close()