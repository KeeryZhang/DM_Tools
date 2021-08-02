#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pprint
import sys
from copy import deepcopy

sys.path.append(".")

if __name__ == "__main__":
    rave_manual_Codelist = r'source\rave_manual_Codelist.xlsx'

    wb = openpyxl.load_workbook(rave_manual_Codelist)
    ws = wb[r'spec_005']

    Codelist_map = {}
    Codelist_map_plus = {}
    Codelist_map_minus = {}
    Codelist_map_nosymbol = {}

    for row in range(2,ws.max_row+1):
        check = ws['A'+str(row)].value

        plus_raw = ws['B'+str(row)].value
        if plus_raw == None:
            pass
        else:
            if ',' in plus_raw:
                plus = plus_raw.split(',')
            else:
                plus = [plus_raw]
            Codelist_map.setdefault((check, '+'), deepcopy(plus))
            Codelist_map_plus.setdefault(check, deepcopy(plus))
            Codelist_map_nosymbol.setdefault(check, deepcopy(plus))

        minus_raw = ws['C'+str(row)].value
        if minus_raw == None:
            pass
        else:
            if ',' in minus_raw:
                minus = minus_raw.split(',')
            else:
                minus = [minus_raw]
            Codelist_map.setdefault((check, '-'), minus)     
            Codelist_map_minus.setdefault(check, minus)
            if check in Codelist_map_nosymbol:
                Codelist_map_nosymbol[check].extend(minus)
            else:
                Codelist_map_nosymbol.setdefault(check, minus)  

    wb.close()
    
    with open("outs\Codelist_map.py", 'w+') as f:
        f.write('Codelist_map = ' + pprint.pformat(Codelist_map) + '\n\n')
        f.write('Codelist_map_plus = ' + pprint.pformat(Codelist_map_plus) + '\n\n')
        f.write('Codelist_map_minus = ' + pprint.pformat(Codelist_map_minus) + '\n\n')
        f.write('Codelist_map_nosymbol = ' + pprint.pformat(Codelist_map_nosymbol) + '\n\n')
        f.close()
    