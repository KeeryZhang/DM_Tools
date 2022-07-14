#!/usr/bin/python
# -*- coding:UTF-8 -*-

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import os
import sys
import argparse
import copy

M2m = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}

def parse_dmy(s, cut=' '):
    if s == None:
        s = cut.join(['UN', 'UNK', '0000'])
    day_s, mon_s, year_s=s.split(cut)
    mon_s = mon_s.upper()
    if day_s in ['UN', 'UK']:
        day_s = '1'
    if mon_s == 'UNK':
        mon_s = 'JAN'
    if year_s == '0000':
        year_s = '1970'
    return datetime(int(year_s),int(M2m[mon_s]),int(day_s))


def mark(ws, col, row, msg):
    ws[col+str(row)] = msg
    return


def findkeyscolumn(ws, keys_raw, row_range=[1]):
    keys = {}
    keyslist = copy.deepcopy(keys_raw)
    for row in row_range:
        for column in range(1, ws.max_column+1):
            row_letter = get_column_letter(column)
            if ws[row_letter+str(row)].value == None:
                continue
            for key in keyslist:
                if key in ws[row_letter+str(row)].value:
                    keys.setdefault(key, row_letter)
                    keyslist.remove(key)
                    break
    
    for key in keys:
        if keys[key] == None:
            keys.pop(key)
            
    return keys


def exist(checkitem, container):
    if checkitem in container:
        return True
    else:
        return False


def message(msg, rsg):
    if msg == '':
        msg = rsg
    else:
        if rsg != "":
            msg = '\n'.join([msg, rsg])
        else:
            msg = msg
    return msg


def get_files():
    files_raw = os.listdir(SHEETS_PATH)
    files = deepcopy(files_raw)
    for file in files_raw:
        if "checkout" in file:
            files.remove(file)
    return files

def get_a_file(files, filename):
    for file in files:
        if filename in file:
            return file